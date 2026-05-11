"""
qPCR结果处理器 - 完整版GUI (带标签页)
功能：格式转换、ΔΔCt计算、板式转换、数据验证

v2.4 增强：
  • 多内参基因（几何平均，自动 fallback 单内参）
  • 统计检验（Student t / Welch / Mann-Whitney）+ Benjamini-Hochberg FDR
  • 汇总表自动加 n / SEM / pvalue / fdr / 显著性 列
  • Fold-change 柱状图：自动加显著性标星 + 支持 PNG / PDF / SVG 输出
  • Excel 导出：可勾选「嵌入柱状图」把 PNG 写到独立 Sheet
  • 「分析预设」：内参/对照组/统计方法/嵌图开关存为 JSON，可一键回填
"""
import json
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import re
import math
import tempfile
import statistics
from pathlib import Path
from typing import List, Dict, Optional, Sequence, Tuple, Iterable, Mapping
from datetime import datetime
import pandas as pd
import numpy as np

try:
    import yaml  # 配置文件加载（可选依赖）
except ImportError:  # 老环境无 yaml 也能跑
    yaml = None  # type: ignore[assignment]

try:
    from scipy import stats as _scipy_stats  # 统计检验（可选）
except ImportError:  # noqa
    _scipy_stats = None  # type: ignore[assignment]

# .ixo 解析器：支持包导入 (`python -m src.complete_gui`) 与脚本导入两种方式。
try:
    from .ixo_parser import IxoParser, is_ixo_file
except ImportError:  # 当 complete_gui.py 被作为脚本直接执行时
    from ixo_parser import IxoParser, is_ixo_file  # type: ignore[no-redef]


# ──────────────────────────────────────────────────────────────────────
#  统计 / 显著性 工具（v2.4）
# ──────────────────────────────────────────────────────────────────────
SIG_LEVELS: Tuple[Tuple[float, str], ...] = (
    (0.001, "***"),
    (0.01, "**"),
    (0.05, "*"),
)


def significance_stars(p: float) -> str:
    """把 p 值转成 ``ns / * / ** / ***``。NaN/None 视为 ns。"""
    if p is None or (isinstance(p, float) and math.isnan(p)):
        return "ns"
    for thr, mark in SIG_LEVELS:
        if p < thr:
            return mark
    return "ns"


def perform_two_sample_test(
    a: Iterable[float],
    b: Iterable[float],
    method: str = "ttest",
) -> float:
    """两样本检验，返回双侧 p 值。失败 / scipy 缺失 / 样本不足时返回 ``nan``。

    ``method`` 取值: ``ttest`` / ``welch`` / ``mannwhitney`` (大小写不敏感)。
    """
    aa = [float(x) for x in a if pd.notna(x)]
    bb = [float(x) for x in b if pd.notna(x)]
    if len(aa) < 2 or len(bb) < 2:
        return float("nan")
    if _scipy_stats is None:
        return float("nan")
    m = (method or "ttest").strip().lower()
    try:
        if m in ("welch", "welch t", "welch_t", "welcht"):
            _, p = _scipy_stats.ttest_ind(aa, bb, equal_var=False)
        elif m in ("mannwhitney", "mann-whitney", "mann whitney", "u", "mwu"):
            _, p = _scipy_stats.mannwhitneyu(aa, bb, alternative="two-sided")
        else:  # 默认 Student t
            _, p = _scipy_stats.ttest_ind(aa, bb, equal_var=True)
        return float(p)
    except Exception:
        return float("nan")


def benjamini_hochberg(pvalues: Sequence[float]) -> List[float]:
    """简易 BH FDR 校正。NaN 原样返回。"""
    n_total = len(pvalues)
    if n_total == 0:
        return []
    idx_pairs = [(i, p) for i, p in enumerate(pvalues) if pd.notna(p)]
    if not idx_pairs:
        return [float("nan")] * n_total
    idx_pairs.sort(key=lambda x: x[1])
    n = len(idx_pairs)
    adj: List[Optional[float]] = [None] * n_total
    prev = 1.0
    for rank, (orig_idx, p) in enumerate(reversed(idx_pairs), start=1):
        k = n - rank + 1  # 当前在排序中的位置
        q = p * n / k
        prev = min(prev, q)
        adj[orig_idx] = min(prev, 1.0)
    return [adj[i] if adj[i] is not None else float("nan") for i in range(n_total)]


def geometric_mean_safe(values: Sequence[float]) -> float:
    """几何平均（忽略 NaN / 非正数；空集合返回 NaN）。"""
    cleaned = [float(v) for v in values if pd.notna(v) and float(v) > 0]
    if not cleaned:
        return float("nan")
    if len(cleaned) == 1:
        return cleaned[0]
    try:
        return float(statistics.geometric_mean(cleaned))
    except Exception:
        return float(np.exp(np.mean(np.log(cleaned))))


def _normalise_efficiency(value: object) -> float:
    """把用户输入的扩增效率统一成 1.0~2.0 之间的乘子。

    - 1.0 ~ 2.5（常见为 1.85~2.05）：直接当成 base，原样返回
    - 50 ~ 200（百分比写法）：按 ``base = 1 + pct/100`` 转换
    - 0~1（也兼容用户写 0.95 = 95% 的）：``base = 1 + value``
    - 其它非法值：回退到 2.0
    """
    try:
        v = float(value)
    except (TypeError, ValueError):
        return 2.0
    if not pd.notna(v) or v <= 0:
        return 2.0
    if 1.0 <= v <= 2.5:
        return v
    if 50.0 <= v <= 200.0:
        return 1.0 + v / 100.0
    if 0.0 < v < 1.0:
        return 1.0 + v
    return 2.0


def export_foldchange_bar_figure(
    summary: pd.DataFrame,
    ref_gene: str,
    out_path: str | Path,
    *,
    title: str = "Relative expression (2^-ΔΔCt)",
    genes_filter: Optional[Sequence[str]] = None,
    ctrl_group: Optional[str] = None,
    dpi: int = 300,
    annotate_significance: bool = True,
    fig_format: Optional[str] = None,
    ref_genes: Optional[Sequence[str]] = None,
) -> Tuple[bool, str]:
    """根据「汇总表」绘制分组柱状图（均值 ± SEM），用于论文插图。

    新增（v2.4）:
        • 自动跳过所有内参基因（``ref_gene`` + ``ref_genes`` 列表）。
        • 若汇总表里有 ``pvalue`` / ``显著性`` / ``significance`` 列，柱顶自动加 ``*/**/*** /ns``。
        • ``out_path`` 后缀决定格式（``.png`` / ``.pdf`` / ``.svg``）；亦可显式 ``fig_format`` 强制覆盖。

    需要 ``matplotlib``；若未安装返回 ``(False, 提示信息)``。
    """
    try:
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except ImportError:
        return False, "未安装 matplotlib，请执行: pip install matplotlib"

    fc_col = "2^-ΔΔCt"
    if fc_col not in summary.columns:
        for c in summary.columns:
            if "2^-" in str(c) and "Ct" in str(c):
                fc_col = c
                break
        else:
            return False, f"汇总表中找不到 Fold-change 列（期望含「2^-ΔΔCt」）: {list(summary.columns)}"

    if "Gene" not in summary.columns:
        return False, "汇总表缺少 Gene 列"

    excluded_refs = {str(ref_gene)} if ref_gene else set()
    if ref_genes:
        excluded_refs.update(str(r) for r in ref_genes if r)
    df = summary[~summary["Gene"].astype(str).isin(excluded_refs)].copy()
    if genes_filter:
        allow = {str(g) for g in genes_filter}
        df = df[df["Gene"].astype(str).isin(allow)]
    if df.empty:
        return False, "没有可绘制的目标基因（请检查内参选择与基因列表）"

    has_group = "Group" in df.columns
    if not has_group:
        return False, "汇总表缺少 Group 列，无法按处理组绘图"

    sig_col: Optional[str] = None
    for cand in ("显著性", "significance", "Significance", "sig"):
        if cand in df.columns:
            sig_col = cand
            break
    pval_col: Optional[str] = None
    for cand in ("pvalue", "p_value", "P值"):
        if cand in df.columns:
            pval_col = cand
            break

    genes = sorted(df["Gene"].astype(str).unique())
    n_genes = len(genes)
    ncols = min(3, n_genes)
    nrows = (n_genes + ncols - 1) // ncols

    plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "Arial Unicode MS", "DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False

    fig, axes = plt.subplots(nrows, ncols, figsize=(4.2 * ncols, 3.8 * nrows), squeeze=False)
    fig.suptitle(title, fontsize=14, fontweight="bold")

    for idx, gene in enumerate(genes):
        r, c = divmod(idx, ncols)
        ax = axes[r][c]
        sub = df[df["Gene"] == gene]
        agg = (
            sub.groupby("Group", sort=False)[fc_col]
            .agg(["mean", "count", "std"])
            .reset_index()
        )
        agg["sem"] = agg.apply(
            lambda row: row["std"] / (row["count"] ** 0.5) if row["count"] > 1 and pd.notna(row["std"]) else 0.0,
            axis=1,
        )
        order = list(agg["Group"])
        if ctrl_group and ctrl_group in order:
            order = [ctrl_group] + [g for g in order if g != ctrl_group]
        else:
            order = sorted(order, key=str)
        agg = agg.set_index("Group").reindex(order).reset_index()

        x = range(len(agg))
        bars = ax.bar(
            x,
            agg["mean"],
            yerr=agg["sem"],
            capsize=4,
            color="#6366f1",
            edgecolor="#312e81",
            linewidth=0.8,
            error_kw={"linewidth": 1.2, "ecolor": "#475569"},
        )

        # 柱顶显著性标星（每个非对照组各取一个 sig 标记）。
        if annotate_significance and (sig_col or pval_col):
            for bar, group_name in zip(bars, agg["Group"]):
                if ctrl_group and str(group_name) == str(ctrl_group):
                    continue
                rows = sub[sub["Group"] == group_name]
                mark: Optional[str] = None
                if sig_col and not rows.empty:
                    cand = rows[sig_col].dropna()
                    if not cand.empty:
                        mark = str(cand.iloc[0])
                if not mark and pval_col and not rows.empty:
                    cand = rows[pval_col].dropna()
                    if not cand.empty:
                        mark = significance_stars(float(cand.iloc[0]))
                if not mark or mark.lower() == "nan":
                    continue
                bar_top = bar.get_height()
                row_match = agg[agg["Group"] == group_name]
                err = float(row_match["sem"].iloc[0]) if len(row_match) else 0.0
                y_pos = (bar_top if pd.notna(bar_top) else 0.0) + (err if pd.notna(err) else 0.0)
                y_pos += max(0.05, abs(y_pos) * 0.05)
                ax.text(
                    bar.get_x() + bar.get_width() / 2,
                    y_pos,
                    mark,
                    ha="center", va="bottom", fontsize=11, fontweight="bold",
                    color="#1f2937",
                )
        ax.set_xticks(list(x))
        ax.set_xticklabels(agg["Group"], rotation=25, ha="right")
        ax.set_ylabel("2^-ΔΔCt (mean ± SEM)")
        ax.set_title(str(gene), fontweight="bold")
        ax.axhline(1.0, color="#94a3b8", linestyle="--", linewidth=0.8, label="FC=1")
        ax.grid(axis="y", linestyle=":", alpha=0.5)

    # 隐藏多余子图
    for j in range(n_genes, nrows * ncols):
        r, c = divmod(j, ncols)
        axes[r][c].set_visible(False)

    fig.tight_layout()
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    save_kwargs: Dict[str, object] = {"dpi": dpi, "bbox_inches": "tight", "facecolor": "white"}
    if fig_format:
        save_kwargs["format"] = fig_format.lower().lstrip(".")
    fig.savefig(out, **save_kwargs)
    plt.close(fig)
    return True, str(out.resolve())


def embed_image_into_excel(
    excel_path: str | Path,
    image_path: str | Path,
    sheet_name: str = "Fold-change 图",
    cell: str = "B2",
    note: Optional[str] = None,
) -> Tuple[bool, str]:
    """把已有 PNG 嵌入到 Excel 工作簿的指定 Sheet。

    需要 ``openpyxl + Pillow``；任一缺失返回 (False, 提示)。
    若同名 Sheet 已存在会被覆盖。
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as XLImage  # type: ignore
    except ImportError:
        return False, "openpyxl 未安装，无法嵌入图片"
    try:
        import PIL  # noqa: F401
    except ImportError:
        return False, "Pillow 未安装（openpyxl 嵌图依赖），请执行: pip install Pillow"

    excel_path = Path(excel_path)
    image_path = Path(image_path)
    if not excel_path.exists():
        return False, f"Excel 不存在: {excel_path}"
    if not image_path.exists():
        return False, f"图片不存在: {image_path}"

    try:
        wb = load_workbook(excel_path)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        if note:
            ws["A1"] = note
        img = XLImage(str(image_path))
        ws.add_image(img, cell)
        wb.save(excel_path)
        return True, str(excel_path.resolve())
    except Exception as exc:
        return False, f"嵌入失败: {exc}"


def read_table_csv(filepath: str | Path) -> pd.DataFrame:
    """读取 CSV/TSV：自动尝试编码与分隔符（逗号 / 制表符 / pandas 嗅探）。"""
    path = Path(filepath)
    last_err: Optional[Exception] = None
    for enc in ("utf-8-sig", "utf-8", "gbk", "cp936"):
        for kwargs in (
            {"sep": None, "engine": "python"},
            {"sep": ","},
            {"sep": "\t"},
        ):
            try:
                return pd.read_csv(path, encoding=enc, **kwargs)
            except Exception as exc:
                last_err = exc
                continue
    if last_err:
        raise last_err
    return pd.read_csv(path)


class Theme:
    """现代化主题 - 优化配色"""
    # 主色调 - 优雅的深蓝紫渐变风格
    PRIMARY = "#6366f1"          # Indigo
    PRIMARY_DARK = "#4f46e5"     # 深蓝紫
    PRIMARY_LIGHT = "#818cf8"    # 浅蓝紫
    PRIMARY_HOVER = "#4338ca"    # 悬停色

    # 功能色
    SUCCESS = "#22c55e"          # 清新绿
    SUCCESS_LIGHT = "#dcfce7"
    WARNING = "#f59e0b"          # 暖橙
    WARNING_LIGHT = "#fef3c7"
    ERROR = "#ef4444"            # 柔红
    INFO = "#0ea5e9"             # 天蓝
    INFO_LIGHT = "#e0f2fe"

    # 背景色
    BG_MAIN = "#f8fafc"          # 柔白背景
    BG_CARD = "#ffffff"          # 卡片白
    BG_HOVER = "#f1f5f9"         # 悬停灰
    BG_INPUT = "#f8fafc"         # 输入框背景

    # 文字色
    TEXT_PRIMARY = "#1e293b"     # 深灰黑
    TEXT_SECONDARY = "#64748b"   # 中性灰
    TEXT_MUTED = "#94a3b8"       # 浅灰
    TEXT_WHITE = "#ffffff"

    # 边框
    BORDER = "#e2e8f0"
    BORDER_FOCUS = "#a5b4fc"

    # 字体
    FONT_TITLE = ("Microsoft YaHei UI", 16, "bold")
    FONT_HEADING = ("Microsoft YaHei UI", 12, "bold")
    FONT_NORMAL = ("Microsoft YaHei UI", 10)
    FONT_SMALL = ("Microsoft YaHei UI", 9)
    FONT_BUTTON = ("Microsoft YaHei UI", 10, "bold")


class PCRDataParser:
    """PCR原始数据解析器"""
    ROWS = list('ABCDEFGHIJKLMNOP')
    
    def __init__(self):
        self.data: List[Dict] = []
        self.data_dict: Dict = {}
    
    def parse_file(self, filepath: str) -> List[Dict]:
        """解析 qPCR 原始结果文件。

        自动识别两种 Roche LightCycler 480 导出格式：

        * ``.txt`` —— 软件中 *Export → Cp Values* 输出的制表符文本；
        * ``.ixo`` —— LCS480 的原生实验文件（IXOS 序列化），由
          :class:`src.ixo_parser.IxoParser` 解析，输出与 .txt 完全一致。
        """
        if is_ixo_file(filepath):
            ixo = IxoParser()
            ixo.parse_file(filepath)
            self.data = ixo.data
            self.data_dict = {(d['Row'], d['Col']): d for d in self.data}
            return self.data

        data: List[Dict] = []
        # 优先 utf-8，回退 utf-8-sig / gbk，覆盖国内常见 Windows 中文导出。
        for enc in ('utf-8', 'utf-8-sig', 'gbk', 'cp936'):
            try:
                with open(filepath, 'r', encoding=enc) as f:
                    lines = f.readlines()
                break
            except UnicodeDecodeError:
                continue
        else:
            raise UnicodeDecodeError(
                'utf-8', b'', 0, 1, f'无法识别 {filepath} 的文本编码'
            )
        for line in lines[2:]:
            parts = line.strip().split('\t')
            if len(parts) >= 5:
                pos = parts[2]
                match = re.match(r'([A-P])(\d+)', pos)
                if match:
                    cp_val = parts[4].strip() if parts[4] else ''
                    cp_val = re.sub(r'[<>].*', '', cp_val).strip()
                    data.append({
                        'Pos': pos, 'Name': parts[3], 'Cp': cp_val,
                        'Row': match.group(1), 'Col': int(match.group(2))
                    })
        self.data = data
        self.data_dict = {(d['Row'], d['Col']): d for d in data}
        return data
    
    def export_sample_gene(self, cols_per_sample: int, rows_per_gene: int,
                          sample_names: List[str] = None,
                          gene_names: List[str] = None) -> pd.DataFrame:
        """导出Sample-Gene布局（Sample在行，Gene在列）- 原始数据"""
        if not self.data:
            raise ValueError("没有数据")
        num_samples = 24 // cols_per_sample
        num_genes = 16 // rows_per_gene
        ct_per_sample = cols_per_sample * rows_per_gene

        if not sample_names:
            sample_names = [f"Sample{i+1}" for i in range(num_samples)]
        if not gene_names:
            gene_names = [f"Gene{i+1}" for i in range(num_genes)]

        excel_data = []
        for sample_idx in range(num_samples):
            col_start = sample_idx * cols_per_sample + 1
            s_name = sample_names[sample_idx] if sample_idx < len(sample_names) else f"Sample{sample_idx+1}"
            for ct_idx in range(ct_per_sample):
                row_offset = ct_idx // cols_per_sample
                col_offset = ct_idx % cols_per_sample
                row_data = [s_name]
                for gene_idx in range(num_genes):
                    gene_row_start = gene_idx * rows_per_gene
                    gene_rows = self.ROWS[gene_row_start:gene_row_start + rows_per_gene]
                    current_row = gene_rows[row_offset] if row_offset < len(gene_rows) else ''
                    col = col_start + col_offset
                    key = (current_row, col)
                    cp = self.data_dict.get(key, {}).get('Cp', '')
                    row_data.append(cp)
                excel_data.append(row_data)

        columns = ['Sample'] + [gene_names[i] if i < len(gene_names) else f"Gene{i+1}"
                               for i in range(num_genes)]
        return pd.DataFrame(excel_data, columns=columns)

    def export_sample_gene_mean(self, cols_per_sample: int, rows_per_gene: int,
                                sample_names: List[str] = None,
                                gene_names: List[str] = None) -> pd.DataFrame:
        """导出Sample-Gene布局 - 计算每个样本每个基因的平均CT值"""
        if not self.data:
            raise ValueError("没有数据")
        num_samples = 24 // cols_per_sample
        num_genes = 16 // rows_per_gene
        ct_per_sample = cols_per_sample * rows_per_gene

        if not sample_names:
            sample_names = [f"Sample{i+1}" for i in range(num_samples)]
        if not gene_names:
            gene_names = [f"Gene{i+1}" for i in range(num_genes)]

        # 收集每个样本每个基因的所有CT值
        sample_gene_cts: Dict[tuple, List[float]] = {}

        for sample_idx in range(num_samples):
            col_start = sample_idx * cols_per_sample + 1
            s_name = sample_names[sample_idx] if sample_idx < len(sample_names) else f"Sample{sample_idx+1}"

            for gene_idx in range(num_genes):
                g_name = gene_names[gene_idx] if gene_idx < len(gene_names) else f"Gene{gene_idx+1}"
                key = (s_name, g_name)
                if key not in sample_gene_cts:
                    sample_gene_cts[key] = []

                gene_row_start = gene_idx * rows_per_gene
                gene_rows = self.ROWS[gene_row_start:gene_row_start + rows_per_gene]

                # 收集该样本该基因的所有重复孔的CT值
                for ct_idx in range(ct_per_sample):
                    row_offset = ct_idx // cols_per_sample
                    col_offset = ct_idx % cols_per_sample
                    current_row = gene_rows[row_offset] if row_offset < len(gene_rows) else ''
                    col = col_start + col_offset
                    pos_key = (current_row, col)
                    cp_str = self.data_dict.get(pos_key, {}).get('Cp', '')
                    if cp_str:
                        try:
                            cp_val = float(cp_str)
                            sample_gene_cts[key].append(cp_val)
                        except ValueError:
                            pass

        # 生成平均值表格
        excel_data = []
        for sample_idx in range(num_samples):
            s_name = sample_names[sample_idx] if sample_idx < len(sample_names) else f"Sample{sample_idx+1}"
            row_data = [s_name]
            for gene_idx in range(num_genes):
                g_name = gene_names[gene_idx] if gene_idx < len(gene_names) else f"Gene{gene_idx+1}"
                key = (s_name, g_name)
                ct_list = sample_gene_cts.get(key, [])
                if ct_list:
                    mean_ct = np.mean(ct_list)
                    row_data.append(round(mean_ct, 4))
                else:
                    row_data.append('')
            excel_data.append(row_data)

        columns = ['Sample'] + [gene_names[i] if i < len(gene_names) else f"Gene{i+1}"
                               for i in range(num_genes)]
        return pd.DataFrame(excel_data, columns=columns)

    def export_long_format_with_mean(self, cols_per_sample: int, rows_per_gene: int,
                                     sample_names: List[str] = None,
                                     gene_names: List[str] = None,
                                     samples_per_group: int = 1,
                                     group_names: List[str] = None) -> pd.DataFrame:
        """导出长格式数据（适合ΔΔCt计算）- 包含Sample, Group, Gene, CT均值"""
        mean_df = self.export_sample_gene_mean(cols_per_sample, rows_per_gene,
                                               sample_names, gene_names)

        num_samples = len(mean_df)
        num_groups = (num_samples + samples_per_group - 1) // samples_per_group

        # 生成组名
        if not group_names:
            group_names = [f"Group{i+1}" for i in range(num_groups)]

        # 将宽格式转换为长格式，添加Group信息
        long_data = []
        for idx, row in mean_df.iterrows():
            sample = row['Sample']
            group_idx = idx // samples_per_group
            group_name = group_names[group_idx] if group_idx < len(group_names) else f"Group{group_idx+1}"

            for col in mean_df.columns[1:]:  # 跳过Sample列
                ct_val = row[col]
                if ct_val != '' and pd.notna(ct_val):
                    long_data.append({
                        'Sample': sample,
                        'Group': group_name,
                        'Gene': col,
                        'Ct': ct_val
                    })
        return pd.DataFrame(long_data)


class DeltaCtCalculator:
    """ΔΔCt计算器 - 输出完整分析表格（包含所有中间步骤）

    v2.4 增强：
        • ``ref_genes``：可传 list[str] 多内参基因；引擎按 *几何平均* 计算每组的综合内参 Ct。
          只传 ``ref_gene`` 仍然是单内参（与旧版本 100% 兼容）。
        • ``stat_test``：``ttest`` / ``welch`` / ``mannwhitney``。汇总表自动加 ``n / SEM /
          pvalue / fdr / 显著性`` 列。
    """

    def __init__(self):
        self.raw_data: pd.DataFrame = None
        self.results: pd.DataFrame = None
        self.summary: pd.DataFrame = None  # 汇总表
        self.side_by_side: pd.DataFrame = None  # 并排格式（内参-目的基因对比）
        self.available_genes: List[str] = []
        self.available_samples: List[str] = []
        # v2.4 新增：本次运行的元数据
        self.last_ref_genes: List[str] = []
        self.last_stat_test: str = "ttest"
        # v2.5 新增：定量方法 + 基因效率
        self.last_method: str = "ddct"
        self.last_gene_efficiencies: Dict[str, float] = {}

    def _fold_change_for(self, gene: str, delta_delta_ct: float) -> float:
        """根据 ``last_method`` 与 ``last_gene_efficiencies`` 决定 fold-change 公式。

        - ``ddct`` (默认): ``2^-ΔΔCt``
        - ``pfaffl``: ``E_gene^-ΔΔCt``，``E_gene`` 缺失时回退到 2.0（与 ddct 等价）
        """
        if pd.isna(delta_delta_ct):
            return np.nan
        if (self.last_method or "ddct").lower() == "pfaffl":
            base = float(self.last_gene_efficiencies.get(str(gene), 2.0))
            if base <= 1.0 or not pd.notna(base):
                base = 2.0
        else:
            base = 2.0
        try:
            return float(base) ** float(-delta_delta_ct)
        except (OverflowError, ValueError):
            return np.nan

    def load_data(self, filepath: str) -> pd.DataFrame:
        """加载数据"""
        ext = Path(filepath).suffix.lower()
        if ext in ['.xlsx', '.xls']:
            df = pd.read_excel(filepath)
        elif ext == '.csv':
            df = pd.read_csv(filepath)
        else:
            df = pd.read_csv(filepath, sep='\t')
        self.raw_data = df
        return df

    def detect_genes_and_samples(self, sample_col: str, gene_col: str):
        """检测数据中的基因和样本列表"""
        if self.raw_data is None:
            return [], []
        self.available_genes = self.raw_data[gene_col].dropna().unique().tolist()
        self.available_samples = self.raw_data[sample_col].dropna().unique().tolist()
        return self.available_genes, self.available_samples

    def calculate(self, sample_col: str, gene_col: str, ct_col: str,
                 ref_gene: str, ctrl_sample: str, group_col: str = None,
                 ref_genes: Optional[Sequence[str]] = None,
                 stat_test: str = "ttest",
                 method: str = "ddct",
                 gene_efficiencies: Optional[Mapping[str, float]] = None,
                 ) -> pd.DataFrame:
        """
        完整ΔΔCt计算 - 支持按组计算内参均值

        参数:
        - group_col: 分组列名（可选）。如果提供，则使用组内参均值；否则使用样本内参均值
        - ref_genes: 多内参列表；提供时取 *几何平均*。若为空 / None 则退化为单内参 ``ref_gene``。
        - stat_test: 统计检验方法 ``ttest`` / ``welch`` / ``mannwhitney``。
        - method: ``ddct`` (默认 2^-ΔΔCt) 或 ``pfaffl`` (Pfaffl 法，按 ``E_gene^-ΔΔCt`` 计算
          fold-change)。当 ``method='pfaffl'`` 而某基因没在 ``gene_efficiencies`` 里指定，
          回退到 ``E=2.0`` (即与 ΔΔCt 等价)。
        - gene_efficiencies: ``{基因名: efficiency}`` 字典。``efficiency`` 是扩增效率
          系数（理论值 2.0 = 100% 高效；典型 1.85~2.05）。也接受百分比写法 ``efficiency=100``
          (会被自动除以 50 转成 2.0)；或直接 ``1.95`` 这种 1~2 的浮点。

        输出列:
        - Sample Name: 样本名
        - Group: 处理组（如有）
        - Target Name: 基因名
        - CT Value: 原始Ct值
        - 组内参Ct均值: 组内所有样本的内参基因Ct平均值（用于计算ΔCt）
        - 样本目的基因Ct均值: 每个样本每个目的基因的Ct平均值
        - ΔCt: 目的基因Ct均值 - 组内参Ct均值
        - 对照组ΔCt均值: 对照组的ΔCt平均值
        - ΔΔCt: 样本ΔCt - 对照组ΔCt均值
        - 2^-ΔΔCt: 相对表达量（method=ddct 时）/ E^-ΔΔCt（method=pfaffl 时）
        """
        df = self.raw_data.copy()
        df[ct_col] = pd.to_numeric(df[ct_col], errors='coerce')

        # ===== 多内参规整 =====
        ref_list: List[str] = []
        if ref_genes:
            ref_list = [str(g).strip() for g in ref_genes if str(g).strip()]
        if not ref_list and ref_gene:
            ref_list = [str(ref_gene).strip()]
        ref_list = list(dict.fromkeys(ref_list))  # 去重保序
        if not ref_list:
            raise ValueError("必须指定至少一个内参基因 (ref_gene 或 ref_genes)")
        primary_ref = ref_list[0]
        ref_set = set(ref_list)
        self.last_ref_genes = ref_list
        self.last_stat_test = (stat_test or "ttest").strip().lower()
        self.last_method = (method or "ddct").strip().lower()
        self.last_gene_efficiencies: Dict[str, float] = {}
        if gene_efficiencies:
            for gene_name, eff in gene_efficiencies.items():
                self.last_gene_efficiencies[str(gene_name)] = _normalise_efficiency(eff)

        samples = df[sample_col].unique()
        genes = df[gene_col].unique()
        target_genes = [g for g in genes if g not in ref_set]

        # 检查是否有Group列
        has_group = group_col and group_col in df.columns

        if has_group:
            groups = df[group_col].unique()
            # 建立样本到组的映射
            sample_to_group = {}
            for _, row in df[[sample_col, group_col]].drop_duplicates().iterrows():
                sample_to_group[row[sample_col]] = row[group_col]
        else:
            # 没有分组，每个样本自成一组
            sample_to_group = {s: s for s in samples}
            groups = samples

        # ===== 1. 计算每个组的内参Ct均值（多内参→几何平均） =====
        # 对单内参等价于普通算术均值，对多内参先按基因均值再几何平均，避免不同量级
        # 内参基因互相挤压。
        group_ref_ct_mean = {}
        for group in groups:
            if has_group:
                group_samples = df[df[group_col] == group][sample_col].unique()
                base = df[(df[sample_col].isin(group_samples)) & (df[gene_col].isin(ref_set))]
            else:
                base = df[(df[sample_col] == group) & (df[gene_col].isin(ref_set))]
            if base.empty:
                group_ref_ct_mean[group] = np.nan
                continue
            per_gene_means = base.groupby(gene_col)[ct_col].mean().dropna()
            if per_gene_means.empty:
                group_ref_ct_mean[group] = np.nan
                continue
            if len(ref_list) == 1:
                group_ref_ct_mean[group] = float(per_gene_means.iloc[0])
            else:
                group_ref_ct_mean[group] = geometric_mean_safe(per_gene_means.tolist())

        # ===== 2. 计算每个样本每个基因的Ct均值 =====
        sample_gene_ct_mean = {}
        for sample in samples:
            for gene in genes:
                data = df[(df[sample_col] == sample) & (df[gene_col] == gene)][ct_col]
                sample_gene_ct_mean[(sample, gene)] = data.mean() if len(data) > 0 else np.nan

        # ===== 3. 计算ΔCt（使用组内参均值）=====
        delta_ct_values = {}
        for sample in samples:
            group = sample_to_group.get(sample, sample)
            group_ref_mean = group_ref_ct_mean.get(group, np.nan)
            for gene in target_genes:
                target_mean = sample_gene_ct_mean.get((sample, gene), np.nan)
                delta_ct_values[(sample, gene)] = target_mean - group_ref_mean

        # ===== 4. 计算对照组ΔCt均值 =====
        # 找到对照组（可能是组名或样本名）
        ctrl_delta_ct_mean = {}
        for gene in target_genes:
            ctrl_deltas = []
            for sample in samples:
                group = sample_to_group.get(sample, sample)
                # 如果是对照组的样本
                if group == ctrl_sample or sample == ctrl_sample:
                    dct = delta_ct_values.get((sample, gene), np.nan)
                    if pd.notna(dct):
                        ctrl_deltas.append(dct)
            ctrl_delta_ct_mean[gene] = np.mean(ctrl_deltas) if ctrl_deltas else np.nan

        # ===== 5. 构建结果表格 =====
        results = []
        for sample in samples:
            group = sample_to_group.get(sample, sample)
            group_ref_mean = group_ref_ct_mean.get(group, np.nan)

            for gene in genes:
                gene_data = df[(df[sample_col] == sample) & (df[gene_col] == gene)]
                ct_values = gene_data[ct_col].dropna().tolist()
                sample_gene_mean = sample_gene_ct_mean.get((sample, gene), np.nan)

                # 内参基因行（多内参时所有内参基因都用此分支）
                if gene in ref_set:
                    for i, ct in enumerate(ct_values):
                        row = {
                            'Sample Name': sample,
                            'Target Name': gene,
                            'CT Value': ct,
                            '组内参Ct均值': group_ref_mean if i == 0 else None,
                            '样本目的基因Ct均值': None,
                            'ΔCt': None,
                            '对照组ΔCt均值': None,
                            'ΔΔCt': None,
                            '2^-ΔΔCt': None,
                        }
                        if has_group:
                            row['Group'] = group
                        results.append(row)
                else:
                    # 目的基因行
                    delta_ct = delta_ct_values.get((sample, gene), np.nan)
                    ctrl_dct = ctrl_delta_ct_mean.get(gene, np.nan)
                    delta_delta_ct = delta_ct - ctrl_dct if pd.notna(delta_ct) and pd.notna(ctrl_dct) else np.nan
                    fold_change = self._fold_change_for(gene, delta_delta_ct)

                    for i, ct in enumerate(ct_values):
                        row = {
                            'Sample Name': sample,
                            'Target Name': gene,
                            'CT Value': ct,
                            '组内参Ct均值': group_ref_mean if i == 0 else None,
                            '样本目的基因Ct均值': sample_gene_mean if i == 0 else None,
                            'ΔCt': delta_ct if i == 0 else None,
                            '对照组ΔCt均值': ctrl_dct if i == 0 else None,
                            'ΔΔCt': delta_delta_ct if i == 0 else None,
                            '2^-ΔΔCt': fold_change if i == 0 else None,
                        }
                        if has_group:
                            row['Group'] = group
                        results.append(row)

        result_df = pd.DataFrame(results)

        if has_group:
            output_cols = ['Sample Name', 'Group', 'Target Name', 'CT Value',
                          '组内参Ct均值', '样本目的基因Ct均值',
                          'ΔCt', '对照组ΔCt均值', 'ΔΔCt', '2^-ΔΔCt']
        else:
            output_cols = ['Sample Name', 'Target Name', 'CT Value',
                          '组内参Ct均值', '样本目的基因Ct均值',
                          'ΔCt', '对照组ΔCt均值', 'ΔΔCt', '2^-ΔΔCt']
        self.results = result_df[[c for c in output_cols if c in result_df.columns]]

        # ===== 6. 生成汇总表（per-sample × gene）+ 统计检验 =====
        summary_data = []
        for sample in samples:
            group = sample_to_group.get(sample, sample)
            group_ref_mean = group_ref_ct_mean.get(group, np.nan)

            for gene in target_genes:
                delta_ct = delta_ct_values.get((sample, gene), np.nan)
                ctrl_dct = ctrl_delta_ct_mean.get(gene, np.nan)
                delta_delta_ct = delta_ct - ctrl_dct if pd.notna(delta_ct) and pd.notna(ctrl_dct) else np.nan
                fold_change = self._fold_change_for(gene, delta_delta_ct)

                row = {
                    'Sample': sample,
                    'Gene': gene,
                    '组内参Ct均值': group_ref_mean,
                    '样本目的基因Ct均值': sample_gene_ct_mean.get((sample, gene), np.nan),
                    'ΔCt': delta_ct,
                    '对照组ΔCt均值': ctrl_dct,
                    'ΔΔCt': delta_delta_ct,
                    '2^-ΔΔCt': fold_change
                }
                if has_group:
                    row['Group'] = group
                summary_data.append(row)

        self.summary = pd.DataFrame(summary_data)
        self._augment_summary_with_stats(
            target_genes, samples, sample_to_group,
            delta_ct_values, ctrl_delta_ct_mean, str(ctrl_sample),
        )

        # ===== 7. 生成并排格式（内参-目的基因对比）=====
        self._generate_side_by_side(
            df, sample_col, gene_col, ct_col, primary_ref, target_genes,
            samples, sample_to_group, group_ref_ct_mean, sample_gene_ct_mean,
            delta_ct_values, ctrl_delta_ct_mean, has_group
        )

        return self.results

    def _augment_summary_with_stats(
        self,
        target_genes: Sequence[str],
        samples: Sequence,
        sample_to_group: Mapping,
        delta_ct_values: Mapping,
        ctrl_delta_ct_mean: Mapping,
        ctrl_group: str,
    ) -> None:
        """给 ``self.summary`` 追加 ``n / fc_sem / pvalue / fdr / 显著性`` 列。

        策略：对每个 (gene, group) 用对照组 ΔCt vs 该处理组 ΔCt 做两样本检验；同 (gene, group)
        在 per-sample 行里共享统计结果，画图函数即可直接读出标星。
        """
        if self.summary is None or self.summary.empty:
            return
        summ = self.summary
        if "Group" not in summ.columns:
            return
        method = self.last_stat_test or "ttest"

        # 收集每基因 ΔCt by group
        group_dct: Dict[Tuple[str, str], List[float]] = {}
        for sample in samples:
            grp = str(sample_to_group.get(sample, sample))
            for gene in target_genes:
                v = delta_ct_values.get((sample, gene), np.nan)
                if pd.notna(v):
                    group_dct.setdefault((str(gene), grp), []).append(float(v))

        pair_keys: List[Tuple[str, str]] = []
        pvals: List[float] = []
        n_map: Dict[Tuple[str, str], int] = {}
        fc_means: Dict[Tuple[str, str], float] = {}
        fc_sems: Dict[Tuple[str, str], float] = {}

        # 计算 (gene, group) → fc 均值/SEM/n（基于 summary 已经填好的 2^-ΔΔCt 列）
        fc_col = "2^-ΔΔCt"
        if fc_col in summ.columns:
            agg = summ.groupby(["Gene", "Group"], dropna=False)[fc_col].agg(["mean", "count", "std"]).reset_index()
            for _, r in agg.iterrows():
                key = (str(r["Gene"]), str(r["Group"]))
                cnt = int(r["count"]) if pd.notna(r["count"]) else 0
                n_map[key] = cnt
                fc_means[key] = float(r["mean"]) if pd.notna(r["mean"]) else float("nan")
                if cnt > 1 and pd.notna(r["std"]):
                    fc_sems[key] = float(r["std"]) / (cnt ** 0.5)
                else:
                    fc_sems[key] = float("nan")

        # 检验：对每个 (gene, group)（非对照）做对照组 vs 该组的 ΔCt 检验
        gene_group_pairs: List[Tuple[str, str]] = []
        for (gene, grp) in group_dct.keys():
            if str(grp) == str(ctrl_group):
                continue
            ctrl_vals = group_dct.get((gene, str(ctrl_group)), [])
            treat_vals = group_dct.get((gene, grp), [])
            p = perform_two_sample_test(ctrl_vals, treat_vals, method=method)
            gene_group_pairs.append((str(gene), str(grp)))
            pvals.append(p)

        adj = benjamini_hochberg(pvals)

        p_map: Dict[Tuple[str, str], float] = {}
        q_map: Dict[Tuple[str, str], float] = {}
        for (g, gr), p, q in zip(gene_group_pairs, pvals, adj):
            p_map[(g, gr)] = p
            q_map[(g, gr)] = q

        def _row_n(r):
            return n_map.get((str(r.get("Gene", "")), str(r.get("Group", ""))), 0)
        def _row_sem(r):
            return fc_sems.get((str(r.get("Gene", "")), str(r.get("Group", ""))), float("nan"))
        def _row_p(r):
            return p_map.get((str(r.get("Gene", "")), str(r.get("Group", ""))), float("nan"))
        def _row_q(r):
            return q_map.get((str(r.get("Gene", "")), str(r.get("Group", ""))), float("nan"))

        summ = summ.copy()
        summ["n"] = summ.apply(_row_n, axis=1)
        summ["fc_sem"] = summ.apply(_row_sem, axis=1)
        summ["pvalue"] = summ.apply(_row_p, axis=1)
        summ["fdr"] = summ.apply(_row_q, axis=1)
        summ["显著性"] = summ["pvalue"].apply(significance_stars)
        # 对照组本身没有 vs 自己的检验，把对照组行的 sig 重置为空
        if "Group" in summ.columns:
            mask_ctrl = summ["Group"].astype(str) == str(ctrl_group)
            summ.loc[mask_ctrl, ["pvalue", "fdr"]] = np.nan
            summ.loc[mask_ctrl, "显著性"] = ""
        self.summary = summ

    def _generate_side_by_side(self, df, sample_col, gene_col, ct_col, ref_gene,
                               target_genes, samples, sample_to_group, group_ref_ct_mean,
                               sample_gene_ct_mean, delta_ct_values, ctrl_delta_ct_mean, has_group):
        """
        生成并排格式的结果表（内参基因列 | 目的基因列）
        格式类似于：
        内参基因                    |  目的基因
        Sample, Gene, CT, 均值...  |  Sample, Gene, CT, ΔCt, ΔΔCt, 2^-ΔΔCt
        """
        # 对于每个目的基因生成一个并排表
        all_side_by_side = []

        for target_gene in target_genes:
            side_data = []

            for sample in samples:
                group = sample_to_group.get(sample, sample)
                group_ref_mean = group_ref_ct_mean.get(group, np.nan)
                target_mean = sample_gene_ct_mean.get((sample, target_gene), np.nan)
                delta_ct = delta_ct_values.get((sample, target_gene), np.nan)
                ctrl_dct = ctrl_delta_ct_mean.get(target_gene, np.nan)
                delta_delta_ct = delta_ct - ctrl_dct if pd.notna(delta_ct) and pd.notna(ctrl_dct) else np.nan
                fold_change = self._fold_change_for(target_gene, delta_delta_ct)

                # 获取内参基因的CT值列表
                ref_cts = df[(df[sample_col] == sample) & (df[gene_col] == ref_gene)][ct_col].dropna().tolist()
                # 获取目的基因的CT值列表
                target_cts = df[(df[sample_col] == sample) & (df[gene_col] == target_gene)][ct_col].dropna().tolist()

                max_reps = max(len(ref_cts), len(target_cts), 1)

                for i in range(max_reps):
                    row = {
                        # 内参基因部分
                        'Sample Name': sample if i == 0 else '',
                        'Target Name': ref_gene if i == 0 else '',
                        'CT Value': ref_cts[i] if i < len(ref_cts) else '',
                        '每个样本内参Ct值均数': group_ref_mean if i == 0 else '',
                        '内参总体Ct值均数': np.mean(list(group_ref_ct_mean.values())) if i == 0 else '',
                        # 分隔列
                        '': '',
                        # 目的基因部分
                        'Sample Name.1': sample if i == 0 else '',
                        'Target Name.1': target_gene if i == 0 else '',
                        'CT Value.1': target_cts[i] if i < len(target_cts) else '',
                        '每个样本目的基因Ct值均数': target_mean if i == 0 else '',
                        '△Ct=目的基因CT-内参CT均值': delta_ct if i == 0 else '',
                        '平均值': np.mean([delta_ct_values.get((s, target_gene), np.nan)
                                          for s in samples if pd.notna(delta_ct_values.get((s, target_gene), np.nan))]) if i == 0 else '',
                        '△Ct=目的基因△CT-内参△CT均值': delta_delta_ct if i == 0 else '',
                        '2^-△△Ct': fold_change if i == 0 else '',
                    }
                    side_data.append(row)

            all_side_by_side.extend(side_data)
            # 添加空行分隔不同基因
            all_side_by_side.append({k: '' for k in side_data[0].keys()} if side_data else {})

        self.side_by_side = pd.DataFrame(all_side_by_side)


class ModernButton(tk.Canvas):
    """现代化按钮 - 带悬停效果和圆角"""

    def __init__(self, parent, text="", command=None, bg=Theme.PRIMARY,
                 fg=Theme.TEXT_WHITE, width=120, height=36, **kwargs):
        super().__init__(parent, width=width, height=height,
                        bg=parent.cget('bg'), highlightthickness=0, **kwargs)
        self.command = command
        self.bg_color = bg
        self.fg_color = fg
        self.width = width
        self.height = height
        self.text = text
        self.hover = False

        self._draw()
        self.bind("<Enter>", self._on_enter)
        self.bind("<Leave>", self._on_leave)
        self.bind("<Button-1>", self._on_click)

    def _draw(self):
        self.delete("all")
        r = 8  # 圆角半径
        color = self._lighten_color(self.bg_color, 20) if self.hover else self.bg_color

        # 绘制圆角矩形
        self.create_rounded_rect(2, 2, self.width-2, self.height-2, r, fill=color, outline="")

        # 绘制文字
        self.create_text(self.width/2, self.height/2, text=self.text,
                        fill=self.fg_color, font=Theme.FONT_BUTTON)

    def create_rounded_rect(self, x1, y1, x2, y2, r, **kwargs):
        points = [
            x1+r, y1, x2-r, y1, x2, y1, x2, y1+r,
            x2, y2-r, x2, y2, x2-r, y2, x1+r, y2,
            x1, y2, x1, y2-r, x1, y1+r, x1, y1
        ]
        return self.create_polygon(points, smooth=True, **kwargs)

    def _lighten_color(self, color, percent):
        """提亮颜色"""
        color = color.lstrip('#')
        r, g, b = int(color[0:2], 16), int(color[2:4], 16), int(color[4:6], 16)
        r = min(255, r + int((255 - r) * percent / 100))
        g = min(255, g + int((255 - g) * percent / 100))
        b = min(255, b + int((255 - b) * percent / 100))
        return f"#{r:02x}{g:02x}{b:02x}"

    def _on_enter(self, event):
        self.hover = True
        self.configure(cursor="hand2")
        self._draw()

    def _on_leave(self, event):
        self.hover = False
        self._draw()

    def _on_click(self, event):
        if self.command:
            self.command()


class CompleteGUI:
    """完整版qPCR处理器GUI - 带标签页 (美化版)"""

    def __init__(self):
        self.root = tk.Tk()
        self.root.title("🧬 qPCR 结果处理器 Pro")
        self.root.geometry("1150x850")
        self.root.configure(bg=Theme.BG_MAIN)
        self.root.minsize(900, 700)

        # 居中显示窗口
        self._center_window()

        self.parser = PCRDataParser()
        self.calculator = DeltaCtCalculator()
        # 宽表 → ΔΔCt 标签页有自己独立的状态，避免与 Tab2 互相覆盖。
        self.wide_calculator = DeltaCtCalculator()
        self.wide_long_df: Optional[pd.DataFrame] = None
        self.wide_input_file: Optional[str] = None
        self.wide_groups: List[str] = []
        self.wide_target_vars: Dict[str, tk.BooleanVar] = {}
        self.config = self._load_config()

        self._setup_styles()
        self._create_ui()

    def _center_window(self):
        """居中显示窗口"""
        self.root.update_idletasks()
        w = 1150
        h = 850
        screen_w = self.root.winfo_screenwidth()
        screen_h = self.root.winfo_screenheight()
        x = (screen_w - w) // 2
        y = (screen_h - h) // 2
        self.root.geometry(f"{w}x{h}+{x}+{y}")

    def _load_config(self) -> dict:
        """加载 ``configs/export_formats.yaml``；无文件或 yaml 未安装时返回内置默认。"""
        default = {'layout': {'cols_per_sample': 2, 'rows_per_gene': 2}}
        if yaml is None:
            return default
        config_path = Path(__file__).parent.parent / "configs" / "export_formats.yaml"
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                data = yaml.safe_load(f)
            return data if isinstance(data, dict) else default
        except Exception:
            return default

    def _presets_dir(self) -> Path:
        d = Path(__file__).resolve().parent.parent / "presets"
        d.mkdir(parents=True, exist_ok=True)
        return d

    def _save_layout_preset(self) -> None:
        """将当前格式转换页的布局与名称保存为 JSON 预设。"""
        payload = {
            "version": 1,
            "cols_per_sample": self.cols_var.get(),
            "rows_per_gene": self.rows_var.get(),
            "samples_per_group": self.samples_per_group_var.get(),
            "sample_names": self.sample_names_var.get(),
            "gene_names": self.gene_names_var.get(),
            "group_names": self.group_names_var.get(),
        }
        initial = self._presets_dir() / "my_layout_preset.json"
        path = filedialog.asksaveasfilename(
            title="保存布局预设",
            initialdir=str(self._presets_dir()),
            initialfile=initial.name,
            defaultextension=".json",
            filetypes=[("JSON 预设", "*.json"), ("所有文件", "*.*")],
        )
        if not path:
            return
        try:
            Path(path).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
            self.status_var.set(f"✅ 已保存预设: {Path(path).name}")
            messagebox.showinfo("成功", f"布局预设已保存:\n{path}")
        except Exception as exc:
            messagebox.showerror("错误", f"保存失败: {exc}")

    def _load_layout_preset(self) -> None:
        """从 JSON 预设恢复布局与名称。"""
        path = filedialog.askopenfilename(
            title="加载布局预设",
            initialdir=str(self._presets_dir()),
            filetypes=[("JSON 预设", "*.json"), ("所有文件", "*.*")],
        )
        if not path:
            return
        try:
            data = json.loads(Path(path).read_text(encoding="utf-8"))
            if not isinstance(data, dict):
                raise ValueError("预设格式无效")
            if "cols_per_sample" in data:
                self.cols_var.set(str(data["cols_per_sample"]))
            if "rows_per_gene" in data:
                self.rows_var.set(str(data["rows_per_gene"]))
            if "samples_per_group" in data:
                self.samples_per_group_var.set(str(data["samples_per_group"]))
            if "sample_names" in data and data["sample_names"] is not None:
                self.sample_names_var.set(str(data["sample_names"]))
            if "gene_names" in data and data["gene_names"] is not None:
                self.gene_names_var.set(str(data["gene_names"]))
            if "group_names" in data and data["group_names"] is not None:
                self.group_names_var.set(str(data["group_names"]))
            self.status_var.set(f"✅ 已加载预设: {Path(path).name}")
            messagebox.showinfo("成功", f"已加载布局预设:\n{Path(path).name}")
        except Exception as exc:
            messagebox.showerror("错误", f"加载失败: {exc}")

    def _setup_styles(self):
        """设置样式 - 增强版"""
        style = ttk.Style()
        style.theme_use('clam')

        # 卡片样式
        style.configure("Card.TFrame", background=Theme.BG_CARD)
        style.configure("TLabel", background=Theme.BG_CARD, font=Theme.FONT_NORMAL)

        # 标签页样式 - 更现代
        style.configure("TNotebook", background=Theme.BG_MAIN, borderwidth=0)
        style.configure("TNotebook.Tab",
                       font=Theme.FONT_HEADING,
                       padding=[25, 12],
                       background=Theme.BG_MAIN,
                       foreground=Theme.TEXT_SECONDARY)
        style.map("TNotebook.Tab",
                 background=[("selected", Theme.BG_CARD)],
                 foreground=[("selected", Theme.PRIMARY)],
                 expand=[("selected", [1, 1, 1, 0])])

        # Spinbox样式
        style.configure("TSpinbox",
                       fieldbackground=Theme.BG_INPUT,
                       padding=5)

        # Scrollbar样式
        style.configure("Vertical.TScrollbar",
                       background=Theme.BG_MAIN,
                       troughcolor=Theme.BG_MAIN,
                       borderwidth=0)

        # Treeview样式
        style.configure("Treeview",
                       font=Theme.FONT_NORMAL,
                       rowheight=28,
                       background=Theme.BG_CARD,
                       fieldbackground=Theme.BG_CARD)
        style.configure("Treeview.Heading",
                       font=Theme.FONT_HEADING,
                       background=Theme.PRIMARY,
                       foreground=Theme.TEXT_WHITE)
        style.map("Treeview",
                 background=[("selected", Theme.PRIMARY_LIGHT)],
                 foreground=[("selected", Theme.TEXT_WHITE)])

    def _create_ui(self):
        """创建UI"""
        self._create_header()

        # 标签页容器
        notebook_frame = tk.Frame(self.root, bg=Theme.BG_MAIN)
        notebook_frame.pack(fill=tk.BOTH, expand=True, padx=25, pady=(15, 10))

        self.notebook = ttk.Notebook(notebook_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)

        # 创建各标签页
        self._create_convert_tab()
        self._create_deltact_tab()
        self._create_wide_ddct_tab()

        self._create_statusbar()

    def _create_header(self):
        """创建顶部标题 - 渐变风格"""
        header = tk.Frame(self.root, bg=Theme.PRIMARY_DARK, height=80)
        header.pack(fill=tk.X)
        header.pack_propagate(False)

        # 左侧标题区
        title_frame = tk.Frame(header, bg=Theme.PRIMARY_DARK)
        title_frame.pack(side=tk.LEFT, padx=30, pady=15)

        # 主标题
        tk.Label(title_frame, text="🧬 qPCR 结果处理器",
                font=("Microsoft YaHei UI", 18, "bold"),
                bg=Theme.PRIMARY_DARK, fg=Theme.TEXT_WHITE).pack(anchor=tk.W)

        # 副标题
        tk.Label(title_frame, text="格式转换  ·  ΔΔCt  ·  宽表  ·  .ixo  ·  预设  ·  Fold-change 图",
                font=Theme.FONT_SMALL, bg=Theme.PRIMARY_DARK,
                fg="#c7d2fe").pack(anchor=tk.W, pady=(5, 0))

        # 右侧版本信息
        tk.Label(header, text="v2.5", font=Theme.FONT_SMALL,
                bg=Theme.PRIMARY_DARK, fg="#a5b4fc").pack(side=tk.RIGHT, padx=30)

    def _create_statusbar(self):
        """状态栏 - 美化版"""
        self.status_var = tk.StringVar(value="✨ 就绪 - 请选择文件开始处理")
        status = tk.Frame(self.root, bg=Theme.BG_CARD, height=40)
        status.pack(fill=tk.X, side=tk.BOTTOM)
        status.pack_propagate(False)

        # 分隔线
        tk.Frame(status, bg=Theme.BORDER, height=1).pack(fill=tk.X)

        # 状态图标和文字
        inner = tk.Frame(status, bg=Theme.BG_CARD)
        inner.pack(fill=tk.BOTH, expand=True)

        tk.Label(inner, text="●", font=("Arial", 8),
                bg=Theme.BG_CARD, fg=Theme.SUCCESS).pack(side=tk.LEFT, padx=(25, 8), pady=10)
        tk.Label(inner, textvariable=self.status_var, font=Theme.FONT_NORMAL,
                bg=Theme.BG_CARD, fg=Theme.TEXT_SECONDARY).pack(side=tk.LEFT, pady=10)

    def _create_card(self, parent, title: str, icon: str = "📋") -> tk.Frame:
        """创建卡片 - 带阴影效果"""
        outer = tk.Frame(parent, bg=Theme.BG_MAIN)
        outer.pack(fill=tk.X, pady=10, padx=25)

        # 卡片主体
        card = tk.Frame(outer, bg=Theme.BG_CARD,
                       highlightbackground=Theme.BORDER, highlightthickness=1)
        card.pack(fill=tk.X, ipadx=20, ipady=15)

        # 标题行
        title_row = tk.Frame(card, bg=Theme.BG_CARD)
        title_row.pack(fill=tk.X, pady=(0, 12))

        tk.Label(title_row, text=f"{icon}  {title}", font=Theme.FONT_HEADING,
                bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY).pack(anchor=tk.W)

        # 分隔线
        tk.Frame(card, bg=Theme.BORDER, height=1).pack(fill=tk.X, pady=(0, 12))

        return card

    # ==================== 格式转换Tab ====================
    def _create_convert_tab(self):
        """格式转换标签页 - 美化版"""
        tab = tk.Frame(self.notebook, bg=Theme.BG_MAIN)
        self.notebook.add(tab, text="  📋 格式转换  ")

        canvas = tk.Canvas(tab, bg=Theme.BG_MAIN, highlightthickness=0)
        scrollbar = ttk.Scrollbar(tab, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg=Theme.BG_MAIN)

        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        # 绑定鼠标滚轮
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # 文件选择卡片
        card1 = self._create_card(scroll_frame, "选择qPCR原始结果文件", "📁")
        file_frame = tk.Frame(card1, bg=Theme.BG_CARD)
        file_frame.pack(fill=tk.X)

        self.convert_file_label = tk.Label(file_frame, text="  📄 点击右侧按钮选择文件...",
                                          font=Theme.FONT_NORMAL, bg=Theme.BG_INPUT,
                                          fg=Theme.TEXT_MUTED, padx=15, pady=12, anchor=tk.W,
                                          relief=tk.FLAT)
        self.convert_file_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # 使用ModernButton
        browse_btn = ModernButton(file_frame, text="📂 浏览文件",
                                  command=self._select_convert_file,
                                  bg=Theme.PRIMARY, width=120, height=40)
        browse_btn.pack(side=tk.RIGHT, padx=(15, 0))

        self.convert_input_file = None

        # 布局设置卡片
        card2 = self._create_card(scroll_frame, "布局设置 (384孔板: 24列 × 16行)", "⚙️")

        layout_frame = tk.Frame(card2, bg=Theme.BG_CARD)
        layout_frame.pack(fill=tk.X, pady=5)

        # 每样本列数
        col_frame = tk.Frame(layout_frame, bg=Theme.BG_CARD)
        col_frame.pack(side=tk.LEFT, padx=(0, 40))
        tk.Label(col_frame, text="每样本列数", font=Theme.FONT_SMALL,
                bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED).pack(anchor=tk.W)
        self.cols_var = tk.StringVar(value=str(self.config.get('layout', {}).get('cols_per_sample', 2)))
        cols_spin = ttk.Spinbox(col_frame, from_=1, to=24, width=8,
                                textvariable=self.cols_var, font=Theme.FONT_NORMAL)
        cols_spin.pack(pady=(5, 0))

        # 每基因行数
        row_frame = tk.Frame(layout_frame, bg=Theme.BG_CARD)
        row_frame.pack(side=tk.LEFT, padx=(0, 40))
        tk.Label(row_frame, text="每基因行数", font=Theme.FONT_SMALL,
                bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED).pack(anchor=tk.W)
        self.rows_var = tk.StringVar(value=str(self.config.get('layout', {}).get('rows_per_gene', 2)))
        rows_spin = ttk.Spinbox(row_frame, from_=1, to=16, width=8,
                                textvariable=self.rows_var, font=Theme.FONT_NORMAL)
        rows_spin.pack(pady=(5, 0))

        # 每组样本数（用于ΔΔCt计算时分组）
        group_frame = tk.Frame(layout_frame, bg=Theme.BG_CARD)
        group_frame.pack(side=tk.LEFT)
        tk.Label(group_frame, text="每组样本数", font=Theme.FONT_SMALL,
                bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED).pack(anchor=tk.W)
        self.samples_per_group_var = tk.StringVar(value="1")
        group_spin = ttk.Spinbox(group_frame, from_=1, to=24, width=8,
                                 textvariable=self.samples_per_group_var, font=Theme.FONT_NORMAL)
        group_spin.pack(pady=(5, 0))

        # 分组说明
        group_info = tk.Label(card2,
                             text="💡 每组样本数：同一处理组有几个生物学重复样本，用于计算组内参均值",
                             font=Theme.FONT_SMALL, bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED)
        group_info.pack(anchor=tk.W, pady=(10, 0))

        # 名称设置卡片
        card3 = self._create_card(scroll_frame, "自定义名称 (可选，逗号分隔)", "🏷️")

        # 样本名输入
        tk.Label(card3, text="样本名称", font=Theme.FONT_SMALL,
                bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED).pack(anchor=tk.W)
        self.sample_names_var = tk.StringVar()
        sample_entry = tk.Entry(card3, textvariable=self.sample_names_var, font=Theme.FONT_NORMAL,
                               relief=tk.FLAT, bg=Theme.BG_INPUT, fg=Theme.TEXT_PRIMARY,
                               insertbackground=Theme.PRIMARY)
        sample_entry.pack(fill=tk.X, pady=(5, 15), ipady=8)

        # 基因名输入
        tk.Label(card3, text="基因名称", font=Theme.FONT_SMALL,
                bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED).pack(anchor=tk.W)
        self.gene_names_var = tk.StringVar()
        gene_entry = tk.Entry(card3, textvariable=self.gene_names_var, font=Theme.FONT_NORMAL,
                             relief=tk.FLAT, bg=Theme.BG_INPUT, fg=Theme.TEXT_PRIMARY,
                             insertbackground=Theme.PRIMARY)
        gene_entry.pack(fill=tk.X, pady=(5, 15), ipady=8)

        # 处理组名输入
        tk.Label(card3, text="处理组名称（可选，逗号分隔，如：Control,Treatment1,Treatment2）", font=Theme.FONT_SMALL,
                bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED).pack(anchor=tk.W)
        self.group_names_var = tk.StringVar()
        group_entry = tk.Entry(card3, textvariable=self.group_names_var, font=Theme.FONT_NORMAL,
                              relief=tk.FLAT, bg=Theme.BG_INPUT, fg=Theme.TEXT_PRIMARY,
                              insertbackground=Theme.PRIMARY)
        group_entry.pack(fill=tk.X, pady=(5, 0), ipady=8)

        # 布局预设卡片（便于重复实验快速恢复参数）
        card_preset = self._create_card(scroll_frame, "布局预设 (JSON)", "💾")
        preset_row = tk.Frame(card_preset, bg=Theme.BG_CARD)
        preset_row.pack(fill=tk.X)
        tk.Label(
            preset_row,
            text="将「每样本列数 / 每基因行数 / 每组样本数 / 三类名称」保存为文件，下次一键恢复。",
            font=Theme.FONT_SMALL, bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED,
            wraplength=880, justify=tk.LEFT,
        ).pack(side=tk.LEFT, fill=tk.X, expand=True)
        ModernButton(
            preset_row, text="💾 保存预设", command=self._save_layout_preset,
            bg=Theme.SUCCESS, width=110, height=36,
        ).pack(side=tk.RIGHT, padx=(8, 0))
        ModernButton(
            preset_row, text="📂 加载预设", command=self._load_layout_preset,
            bg=Theme.INFO, width=110, height=36,
        ).pack(side=tk.RIGHT)

        # 操作按钮卡片
        card4 = self._create_card(scroll_frame, "操作", "🚀")

        # 第一行按钮：预览
        btn_frame1 = tk.Frame(card4, bg=Theme.BG_CARD)
        btn_frame1.pack(fill=tk.X, pady=5)
        ModernButton(btn_frame1, text="👁️ 预览原始", command=self._preview_convert,
                    bg=Theme.INFO, width=110, height=42).pack(side=tk.LEFT, padx=(0, 10))
        ModernButton(btn_frame1, text="📊 预览平均CT", command=self._preview_convert_mean,
                    bg=Theme.WARNING, width=120, height=42).pack(side=tk.LEFT, padx=(0, 10))
        ModernButton(btn_frame1, text="📋 预览长格式(分组)", command=self._preview_convert_long,
                    bg="#9333ea", width=150, height=42).pack(side=tk.LEFT)

        # 第二行按钮：导出
        btn_frame2 = tk.Frame(card4, bg=Theme.BG_CARD)
        btn_frame2.pack(fill=tk.X, pady=(10, 5))
        ModernButton(btn_frame2, text="💾 导出Excel (含3个Sheet)", command=self._export_convert,
                    bg=Theme.SUCCESS, width=200, height=42).pack(side=tk.LEFT, padx=(0, 15))
        ModernButton(btn_frame2, text="📋 复制剪贴板", command=self._copy_convert,
                    bg=Theme.PRIMARY, width=130, height=42).pack(side=tk.LEFT)

        # ===== 一键全流程卡片 (.ixo / .txt → 并排格式 ΔΔCt) =====
        card5 = self._create_card(
            scroll_frame, "🚀 一键全流程：直接生成并排格式 ΔΔCt", "⚡"
        )

        info_text = (
            "选好上方文件（支持 .ixo 直读 / .txt 导出）+ 布局参数 + 自定义名称后，"
            "点击下方按钮即可一步生成与「ΔΔCt 计算」标签页同款的「内参/目的基因并排格式」。\n"
            "弹窗里再选好【内参基因】和【对照组】，输出 Excel 同时含 长格式 / 详细 / 汇总 / 并排 / 参数 5 个 Sheet。"
        )
        tk.Label(
            card5, text=info_text, font=Theme.FONT_SMALL,
            bg=Theme.BG_CARD, fg=Theme.TEXT_SECONDARY,
            justify=tk.LEFT, wraplength=900,
        ).pack(anchor=tk.W, pady=(0, 12))

        one_click_row = tk.Frame(card5, bg=Theme.BG_CARD)
        one_click_row.pack(fill=tk.X)
        ModernButton(
            one_click_row, text="🚀 一键 ΔΔCt 全流程导出", command=self._one_click_ddct,
            bg="#dc2626", width=240, height=46,
        ).pack(side=tk.LEFT, padx=(0, 12))
        ModernButton(
            one_click_row, text="📊 仅导出长格式预览", command=self._preview_convert_long,
            bg=Theme.INFO, width=170, height=46,
        ).pack(side=tk.LEFT)

    def _select_convert_file(self):
        """选择转换文件"""
        filepath = filedialog.askopenfilename(
            title="选择qPCR原始结果文件",
            filetypes=[
                ("Roche 数据文件", "*.txt;*.ixo"),
                ("LightCycler IXO", "*.ixo"),
                ("文本文件", "*.txt"),
                ("所有文件", "*.*"),
            ],
        )
        if filepath:
            self.convert_input_file = filepath
            self.convert_file_label.config(text=f"📄 {Path(filepath).name}",
                                          fg=Theme.TEXT_PRIMARY, bg="#e0f2fe")
            try:
                self.parser.parse_file(filepath)
                self.status_var.set(f"✅ 已加载 {len(self.parser.data)} 个数据点")
            except Exception as e:
                messagebox.showerror("错误", f"解析失败: {e}")

    def _get_convert_df(self):
        """获取转换后的DataFrame"""
        if not self.parser.data:
            messagebox.showwarning("警告", "请先选择文件")
            return None

        cols = int(self.cols_var.get())
        rows = int(self.rows_var.get())

        s = self.sample_names_var.get().strip()
        g = self.gene_names_var.get().strip()
        samples = [x.strip() for x in s.split(',') if x.strip()] if s else None
        genes = [x.strip() for x in g.split(',') if x.strip()] if g else None

        return self.parser.export_sample_gene(cols, rows, samples, genes)

    def _get_convert_params(self):
        """获取转换参数"""
        cols = int(self.cols_var.get())
        rows = int(self.rows_var.get())
        s = self.sample_names_var.get().strip()
        g = self.gene_names_var.get().strip()
        samples = [x.strip() for x in s.split(',') if x.strip()] if s else None
        genes = [x.strip() for x in g.split(',') if x.strip()] if g else None
        return cols, rows, samples, genes

    def _get_group_params(self):
        """获取分组参数"""
        samples_per_group = int(self.samples_per_group_var.get())
        grp = self.group_names_var.get().strip()
        groups = [x.strip() for x in grp.split(',') if x.strip()] if grp else None
        return samples_per_group, groups

    def _get_convert_mean_df(self):
        """获取平均CT值的DataFrame"""
        if not self.parser.data:
            messagebox.showwarning("警告", "请先选择文件")
            return None
        cols, rows, samples, genes = self._get_convert_params()
        return self.parser.export_sample_gene_mean(cols, rows, samples, genes)

    def _get_convert_long_df(self):
        """获取长格式的DataFrame（适合ΔΔCt计算）"""
        if not self.parser.data:
            messagebox.showwarning("警告", "请先选择文件")
            return None
        cols, rows, samples, genes = self._get_convert_params()
        samples_per_group, groups = self._get_group_params()
        return self.parser.export_long_format_with_mean(cols, rows, samples, genes,
                                                        samples_per_group, groups)

    def _preview_convert(self):
        """预览转换结果 - 原始数据"""
        df = self._get_convert_df()
        if df is None:
            return
        self._show_preview_window("格式转换预览 - 原始数据", df)

    def _preview_convert_mean(self):
        """预览平均CT值"""
        df = self._get_convert_mean_df()
        if df is None:
            return
        self._show_preview_window("格式转换预览 - 平均CT值", df)

    def _preview_convert_long(self):
        """预览长格式数据（含分组）"""
        df = self._get_convert_long_df()
        if df is None:
            return
        self._show_preview_window("格式转换预览 - 长格式（含分组）", df)

    def _export_convert(self):
        """导出转换结果 - 包含多个Sheet"""
        if not self.parser.data:
            messagebox.showwarning("警告", "请先选择文件")
            return

        output = filedialog.asksaveasfilename(title="保存", defaultextension=".xlsx",
                                             initialfile="qPCR_converted.xlsx",
                                             filetypes=[("Excel", "*.xlsx")])
        if output:
            cols, rows, samples, genes = self._get_convert_params()
            samples_per_group, groups = self._get_group_params()

            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Sheet 1: 原始数据
                df_raw = self.parser.export_sample_gene(cols, rows, samples, genes)
                df_raw.to_excel(writer, sheet_name='原始数据', index=False)

                # Sheet 2: 平均CT值（宽格式）
                df_mean = self.parser.export_sample_gene_mean(cols, rows, samples, genes)
                df_mean.to_excel(writer, sheet_name='平均CT值', index=False)

                # Sheet 3: 长格式（适合ΔΔCt计算，含分组）
                df_long = self.parser.export_long_format_with_mean(cols, rows, samples, genes,
                                                                   samples_per_group, groups)
                df_long.to_excel(writer, sheet_name='长格式_ΔΔCt用', index=False)

            self.status_var.set(f"✅ 已导出: {Path(output).name}")
            messagebox.showinfo("成功", f"已保存:\n{output}\n\n包含3个Sheet:\n• 原始数据\n• 平均CT值\n• 长格式_ΔΔCt用（含Group列）")

    def _copy_convert(self):
        """复制到剪贴板"""
        df = self._get_convert_df()
        if df is None:
            return
        self.root.clipboard_clear()
        self.root.clipboard_append(df.to_csv(sep='\t', index=False))
        self.status_var.set("📋 已复制到剪贴板")

    # ---------- 一键全流程：.ixo / .txt → 并排格式 ΔΔCt ----------
    def _one_click_ddct(self):
        """从已加载的 Roche 数据 (txt/ixo) 一键到「并排格式 ΔΔCt Excel」。

        v2.4：弹窗增加多内参 / 统计方法 / 嵌图选项；输出 Excel 自动嵌入 Fold-change 图。
        """
        if not self.parser.data:
            messagebox.showwarning(
                "提示",
                "请先在上方选择 .ixo 或 .txt 文件，并设置好布局/样本/基因/分组名。",
            )
            return

        try:
            cols, rows, samples, genes = self._get_convert_params()
            samples_per_group, groups = self._get_group_params()
            long_df = self.parser.export_long_format_with_mean(
                cols, rows, samples, genes, samples_per_group, groups,
            )
        except Exception as exc:
            messagebox.showerror("错误", f"生成长格式失败: {exc}")
            return

        if long_df.empty:
            messagebox.showerror("错误", "长格式为空，请检查布局参数")
            return

        gene_list = long_df['Gene'].dropna().unique().tolist()
        group_list = long_df['Group'].dropna().unique().tolist()

        # 弹窗收集：内参基因 + 对照组 + 输出文件名 + 多内参 + 统计方法 + 嵌图
        chosen = self._ask_one_click_options(gene_list, group_list)
        if chosen is None:
            return
        ref_gene = chosen["ref_gene"]
        ctrl_group = chosen["ctrl_group"]
        sample_info = chosen["sample_info"]
        extra_refs = chosen["extra_refs"]
        stat_method = chosen["stat_test"]
        embed_img = chosen["embed_image"]
        quant_method = chosen.get("quant_method", "ddct")

        output = filedialog.asksaveasfilename(
            title="保存一键 ΔΔCt 结果",
            defaultextension=".xlsx",
            initialfile=(
                f"{datetime.now():%Y%m%d}_{sample_info}_Analysis and processing.xlsx"
                if sample_info else f"{datetime.now():%Y%m%d}_oneclick_ddct.xlsx"
            ),
            filetypes=[("Excel", "*.xlsx")],
        )
        if not output:
            return

        try:
            calc = DeltaCtCalculator()
            calc.raw_data = long_df
            ref_full = list(dict.fromkeys([ref_gene] + extra_refs))
            calc.calculate(
                sample_col='Sample', gene_col='Gene', ct_col='Ct',
                ref_gene=ref_gene, ctrl_sample=ctrl_group, group_col='Group',
                ref_genes=ref_full, stat_test=stat_method,
                method=quant_method,
                gene_efficiencies=(self.dc_gene_efficiencies if quant_method == "pfaffl" else None),
            )
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                long_df.to_excel(writer, sheet_name='长格式_ΔΔCt用', index=False)
                calc.results.to_excel(writer, sheet_name='详细计算结果', index=False)
                if calc.summary is not None:
                    calc.summary.to_excel(writer, sheet_name='汇总表', index=False)
                if calc.side_by_side is not None:
                    calc.side_by_side.to_excel(writer, sheet_name='并排格式_Analysis', index=False)
                params = pd.DataFrame({
                    '参数': [
                        '模式', '源文件', '主内参基因', '多内参 (附加)',
                        '统计方法', '对照组', '样本信息', '是否嵌入图',
                    ],
                    '设置值': [
                        '一键全流程',
                        Path(self.convert_input_file).name if self.convert_input_file else '',
                        ref_gene, ', '.join(extra_refs) if extra_refs else '(无)',
                        stat_method, ctrl_group, sample_info,
                        'Yes' if embed_img else 'No',
                    ],
                })
                params.to_excel(writer, sheet_name='计算参数', index=False)

            embedded = False
            embed_msg = ""
            if embed_img and calc.summary is not None:
                try:
                    with tempfile.TemporaryDirectory() as td:
                        png_path = Path(td) / "oneclick_foldchange.png"
                        ok_p, msg_p = export_foldchange_bar_figure(
                            calc.summary, ref_gene, png_path,
                            title=f"2^-ΔΔCt (ref={ref_gene})",
                            ctrl_group=ctrl_group, ref_genes=extra_refs, fig_format="png",
                        )
                        if ok_p:
                            ok_e, msg_e = embed_image_into_excel(
                                output, png_path, sheet_name="Fold-change 图",
                                note=f"Generated {datetime.now():%Y-%m-%d %H:%M}",
                            )
                            embedded = ok_e
                            embed_msg = "嵌入完成" if ok_e else msg_e
                        else:
                            embed_msg = msg_p
                except Exception as exc:
                    embed_msg = f"嵌图失败: {exc}"

            self.status_var.set(f"✅ 一键全流程完成: {Path(output).name}")
            extra_line = f"\n图嵌入: {embed_msg}" if embed_img else ""
            ref_desc = ref_gene if not extra_refs else f"{ref_gene} + {', '.join(extra_refs)} (几何平均)"
            messagebox.showinfo(
                "成功",
                f"一键 ΔΔCt 完成！\n\n输出文件：\n{output}\n\n"
                f"内参基因: {ref_desc}\n对照组: {ctrl_group}\n统计: {stat_method}\n"
                f"详细结果: {len(calc.results)} 行\n并排格式: {len(calc.side_by_side)} 行"
                + extra_line,
            )
        except Exception as exc:
            messagebox.showerror("错误", f"一键全流程失败: {exc}")

    def _ask_one_click_options(
        self, genes: List[str], groups: List[str],
    ) -> Optional[Dict[str, object]]:
        """弹窗收集：主内参 / 多内参 / 对照组 / 统计方法 / 嵌图 / 样本信息。

        返回 ``{ref_gene, extra_refs, ctrl_group, sample_info, stat_test, embed_image}`` 或 None。
        """
        win = tk.Toplevel(self.root)
        win.title("🚀 一键 ΔΔCt 参数")
        win.configure(bg=Theme.BG_MAIN)
        win.transient(self.root)
        win.grab_set()
        w, h = 540, 500
        x = (win.winfo_screenwidth() - w) // 2
        y = (win.winfo_screenheight() - h) // 2
        win.geometry(f"{w}x{h}+{x}+{y}")
        win.resizable(False, False)

        header = tk.Frame(win, bg=Theme.PRIMARY_DARK, height=50)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        tk.Label(
            header, text="🚀 一键 ΔΔCt 参数", font=Theme.FONT_HEADING,
            bg=Theme.PRIMARY_DARK, fg=Theme.TEXT_WHITE,
        ).pack(side=tk.LEFT, padx=20, pady=12)

        body = tk.Frame(win, bg=Theme.BG_CARD)
        body.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)

        # 自动猜默认值
        common_refs = ['GAPDH', 'Gapdh', 'gapdh', 'ACTB', 'Actb', 'actb',
                       'RPL13a', '18S', 'β-actin', 'HPRT']
        common_ctrls = ['Control', 'control', 'CTRL', 'NC', 'WT', 'wt', 'C']
        ref_default = next((g for g in common_refs if g in genes), genes[0] if genes else "")
        ctrl_default = next((g for g in common_ctrls if g in groups), groups[0] if groups else "")

        ref_var = tk.StringVar(value=ref_default)
        ctrl_var = tk.StringVar(value=ctrl_default)
        info_var = tk.StringVar(value="F1-PFC")
        stat_var = tk.StringVar(value="Student t")
        method_var = tk.StringVar(value="2^-ΔΔCt")
        embed_var = tk.BooleanVar(value=True)

        for label, var, options in [
            ("主内参 (Reference):", ref_var, genes),
            ("对照组 (Control Group):", ctrl_var, groups),
        ]:
            row = tk.Frame(body, bg=Theme.BG_CARD)
            row.pack(fill=tk.X, pady=6)
            tk.Label(row, text=label, font=Theme.FONT_NORMAL,
                     bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY, width=22, anchor=tk.W).pack(side=tk.LEFT)
            ttk.Combobox(
                row, textvariable=var,
                values=options if options else ["（无可选项）"],
                font=Theme.FONT_NORMAL, width=22, state="readonly",
            ).pack(side=tk.LEFT)

        # 多内参 listbox
        multi_row = tk.Frame(body, bg=Theme.BG_CARD)
        multi_row.pack(fill=tk.X, pady=6)
        tk.Label(multi_row, text="多内参 (可选):", font=Theme.FONT_NORMAL,
                 bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY, width=22, anchor=tk.NW).pack(side=tk.LEFT, anchor=tk.N)
        m_lb_frame = tk.Frame(multi_row, bg=Theme.BG_CARD)
        m_lb_frame.pack(side=tk.LEFT)
        m_lb = tk.Listbox(
            m_lb_frame, selectmode=tk.MULTIPLE,
            font=Theme.FONT_NORMAL, height=4, width=20,
            bg=Theme.BG_INPUT, fg=Theme.TEXT_PRIMARY,
            highlightbackground=Theme.BORDER, highlightthickness=1, exportselection=False,
        )
        m_lb.pack(side=tk.LEFT)
        for g in genes:
            if str(g) != ref_default:
                m_lb.insert(tk.END, str(g))
        ttk.Scrollbar(m_lb_frame, orient="vertical", command=m_lb.yview).pack(side=tk.LEFT, fill=tk.Y)

        stat_row = tk.Frame(body, bg=Theme.BG_CARD)
        stat_row.pack(fill=tk.X, pady=6)
        tk.Label(stat_row, text="统计方法:", font=Theme.FONT_NORMAL,
                 bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY, width=22, anchor=tk.W).pack(side=tk.LEFT)
        ttk.Combobox(
            stat_row, textvariable=stat_var, state="readonly",
            font=Theme.FONT_NORMAL, width=22,
            values=["Student t", "Welch t", "Mann-Whitney U"],
        ).pack(side=tk.LEFT)

        method_dlg_row = tk.Frame(body, bg=Theme.BG_CARD)
        method_dlg_row.pack(fill=tk.X, pady=6)
        tk.Label(method_dlg_row, text="定量方法:", font=Theme.FONT_NORMAL,
                 bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY, width=22, anchor=tk.W).pack(side=tk.LEFT)
        ttk.Combobox(
            method_dlg_row, textvariable=method_var, state="readonly",
            font=Theme.FONT_NORMAL, width=22,
            values=["2^-ΔΔCt", "Pfaffl (E^-ΔΔCt)"],
        ).pack(side=tk.LEFT)
        tk.Label(method_dlg_row, text="  Pfaffl 时复用 ΔΔCt 标签页的基因效率表",
                 font=Theme.FONT_SMALL, bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED).pack(side=tk.LEFT)

        embed_row = tk.Frame(body, bg=Theme.BG_CARD)
        embed_row.pack(fill=tk.X, pady=6)
        tk.Label(embed_row, text="Excel 嵌图:", font=Theme.FONT_NORMAL,
                 bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY, width=22, anchor=tk.W).pack(side=tk.LEFT)
        tk.Checkbutton(
            embed_row, text="导出 Excel 时嵌入 Fold-change 柱状图 (PNG)",
            variable=embed_var, bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY,
            font=Theme.FONT_NORMAL, activebackground=Theme.BG_CARD,
        ).pack(side=tk.LEFT)

        info_row = tk.Frame(body, bg=Theme.BG_CARD)
        info_row.pack(fill=tk.X, pady=6)
        tk.Label(info_row, text="样本信息（文件名）:", font=Theme.FONT_NORMAL,
                 bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY, width=22, anchor=tk.W).pack(side=tk.LEFT)
        tk.Entry(info_row, textvariable=info_var, font=Theme.FONT_NORMAL,
                 relief=tk.FLAT, bg=Theme.BG_INPUT, fg=Theme.TEXT_PRIMARY, width=24,
                 insertbackground=Theme.PRIMARY).pack(side=tk.LEFT, ipady=6)

        result: Dict[str, Optional[Dict[str, object]]] = {'val': None}

        def _stat_label_to_key(label: str) -> str:
            ll = label.lower()
            if "welch" in ll:
                return "welch"
            if "mann" in ll or "whitney" in ll:
                return "mannwhitney"
            return "ttest"

        def on_ok():
            r, c = ref_var.get().strip(), ctrl_var.get().strip()
            if not r or not c:
                messagebox.showwarning("提示", "请选择主内参与对照组", parent=win)
                return
            extras = [m_lb.get(i) for i in m_lb.curselection() if m_lb.get(i) != r]
            method_label = method_var.get().strip().lower()
            quant_method = "pfaffl" if "pfaffl" in method_label else "ddct"
            result['val'] = {
                "ref_gene": r,
                "extra_refs": extras,
                "ctrl_group": c,
                "sample_info": info_var.get().strip(),
                "stat_test": _stat_label_to_key(stat_var.get()),
                "embed_image": bool(embed_var.get()),
                "quant_method": quant_method,
            }
            win.destroy()

        def on_cancel():
            win.destroy()

        btn_row = tk.Frame(body, bg=Theme.BG_CARD)
        btn_row.pack(fill=tk.X, pady=(20, 0))
        ModernButton(btn_row, text="✅ 开始一键导出", command=on_ok,
                     bg=Theme.SUCCESS, width=160, height=42).pack(side=tk.LEFT, padx=(0, 12))
        ModernButton(btn_row, text="✗ 取消", command=on_cancel,
                     bg=Theme.TEXT_MUTED, width=100, height=42).pack(side=tk.LEFT)

        win.wait_window()
        return result['val']

    # ==================== ΔΔCt计算Tab ====================
    def _create_deltact_tab(self):
        """ΔΔCt计算标签页 - 完整版，带下拉选择"""
        tab = tk.Frame(self.notebook, bg=Theme.BG_MAIN)
        self.notebook.add(tab, text="  🧮 ΔΔCt计算  ")

        canvas = tk.Canvas(tab, bg=Theme.BG_MAIN, highlightthickness=0)
        scrollbar = ttk.Scrollbar(tab, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg=Theme.BG_MAIN)

        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        # ===== Step 1: 文件选择卡片 =====
        card1 = self._create_card(scroll_frame, "Step 1: 选择Ct数据文件 (Excel/CSV)", "📁")
        file_frame = tk.Frame(card1, bg=Theme.BG_CARD)
        file_frame.pack(fill=tk.X)

        self.deltact_file_label = tk.Label(file_frame, text="  📄 点击右侧按钮选择文件...",
                                          font=Theme.FONT_NORMAL, bg=Theme.BG_INPUT,
                                          fg=Theme.TEXT_MUTED, padx=15, pady=12, anchor=tk.W)
        self.deltact_file_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

        ModernButton(file_frame, text="📂 浏览文件", command=self._select_deltact_file,
                    bg=Theme.PRIMARY, width=120, height=40).pack(side=tk.RIGHT, padx=(15, 0))

        self.deltact_input_file = None

        # Sheet选择行
        sheet_frame = tk.Frame(card1, bg=Theme.BG_CARD)
        sheet_frame.pack(fill=tk.X, pady=(10, 0))

        tk.Label(sheet_frame, text="选择Sheet:", font=Theme.FONT_NORMAL,
                bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY).pack(side=tk.LEFT)

        self.dc_sheet_var = tk.StringVar(value="")
        self.sheet_combo = ttk.Combobox(sheet_frame, textvariable=self.dc_sheet_var,
                                        font=Theme.FONT_NORMAL, width=25, state="readonly")
        self.sheet_combo.pack(side=tk.LEFT, padx=(15, 15))
        self.sheet_combo['values'] = ["请先选择Excel文件"]
        self.sheet_combo.bind("<<ComboboxSelected>>", self._on_sheet_selected)

        tk.Label(sheet_frame, text="← 选择包含长格式数据的Sheet", font=Theme.FONT_SMALL,
                bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED).pack(side=tk.LEFT)

        # ===== Step 2: 列名映射卡片 =====
        card2 = self._create_card(scroll_frame, "Step 2: 列名设置（加载文件后自动检测）", "🔧")

        self.dc_sample_col = tk.StringVar(value="Sample")
        self.dc_group_col = tk.StringVar(value="Group")
        self.dc_gene_col = tk.StringVar(value="Gene")
        self.dc_ct_col = tk.StringVar(value="Ct")

        col_settings = [
            ("样本列名", self.dc_sample_col, "数据中代表样本名称的列"),
            ("分组列名", self.dc_group_col, "数据中代表处理组的列（用于计算组内参均值）"),
            ("基因列名", self.dc_gene_col, "数据中代表基因名称的列"),
            ("Ct值列名", self.dc_ct_col, "数据中代表Ct值的列")
        ]

        for label, var, hint in col_settings:
            row = tk.Frame(card2, bg=Theme.BG_CARD)
            row.pack(fill=tk.X, pady=8)

            label_frame = tk.Frame(row, bg=Theme.BG_CARD, width=100)
            label_frame.pack(side=tk.LEFT)
            label_frame.pack_propagate(False)
            tk.Label(label_frame, text=label, font=Theme.FONT_NORMAL,
                    bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY).pack(anchor=tk.W)

            entry = tk.Entry(row, textvariable=var, font=Theme.FONT_NORMAL, relief=tk.FLAT,
                           bg=Theme.BG_INPUT, fg=Theme.TEXT_PRIMARY, width=18,
                           insertbackground=Theme.PRIMARY)
            entry.pack(side=tk.LEFT, ipady=6, padx=(10, 15))

            tk.Label(row, text=hint, font=Theme.FONT_SMALL,
                    bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED).pack(side=tk.LEFT)

        # 检测按钮
        detect_frame = tk.Frame(card2, bg=Theme.BG_CARD)
        detect_frame.pack(fill=tk.X, pady=(10, 0))
        ModernButton(detect_frame, text="🔍 检测基因和样本", command=self._detect_genes_samples,
                    bg=Theme.INFO, width=150, height=38).pack(side=tk.LEFT)
        self.detect_status_label = tk.Label(detect_frame, text="", font=Theme.FONT_SMALL,
                                           bg=Theme.BG_CARD, fg=Theme.SUCCESS)
        self.detect_status_label.pack(side=tk.LEFT, padx=(15, 0))

        # ===== Step 3: 选择内参基因 =====
        card3 = self._create_card(scroll_frame, "Step 3: 选择内参基因（Reference Gene）", "🧬")

        ref_info = tk.Label(card3, text="内参基因用于归一化，常用: GAPDH, ACTB, 18S rRNA 等",
                           font=Theme.FONT_SMALL, bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED)
        ref_info.pack(anchor=tk.W, pady=(0, 10))

        ref_row = tk.Frame(card3, bg=Theme.BG_CARD)
        ref_row.pack(fill=tk.X)

        tk.Label(ref_row, text="主内参:", font=Theme.FONT_NORMAL,
                bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY).pack(side=tk.LEFT)

        self.dc_ref_gene = tk.StringVar(value="")
        self.ref_gene_combo = ttk.Combobox(ref_row, textvariable=self.dc_ref_gene,
                                           font=Theme.FONT_NORMAL, width=25, state="readonly")
        self.ref_gene_combo.pack(side=tk.LEFT, padx=(15, 20))
        self.ref_gene_combo['values'] = ["请先加载数据并检测"]

        tk.Label(ref_row, text="← 从下拉列表选择或手动输入", font=Theme.FONT_SMALL,
                bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED).pack(side=tk.LEFT)

        # 多内参（可选）— 从基因列表里 Ctrl/Shift 多选；选中后用 *几何平均* 计算综合内参 Ct。
        multi_row = tk.Frame(card3, bg=Theme.BG_CARD)
        multi_row.pack(fill=tk.X, pady=(10, 0))
        tk.Label(multi_row, text="多内参 (可选):", font=Theme.FONT_NORMAL,
                 bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY).pack(side=tk.LEFT, anchor=tk.N)
        listbox_frame = tk.Frame(multi_row, bg=Theme.BG_CARD)
        listbox_frame.pack(side=tk.LEFT, padx=(15, 15))
        self.dc_multi_ref_listbox = tk.Listbox(
            listbox_frame, selectmode=tk.MULTIPLE,
            font=Theme.FONT_NORMAL, height=4, width=22,
            bg=Theme.BG_INPUT, fg=Theme.TEXT_PRIMARY,
            highlightbackground=Theme.BORDER, highlightthickness=1,
            exportselection=False,
        )
        self.dc_multi_ref_listbox.pack(side=tk.LEFT)
        ml_sb = ttk.Scrollbar(listbox_frame, orient="vertical", command=self.dc_multi_ref_listbox.yview)
        ml_sb.pack(side=tk.LEFT, fill=tk.Y)
        self.dc_multi_ref_listbox.configure(yscrollcommand=ml_sb.set)
        tk.Label(
            multi_row,
            text=("← 不勾选 = 仅用主内参；多选时按几何平均合成内参\n"
                  "   选中条目作为「补充」，主内参会自动并入"),
            font=Theme.FONT_SMALL, bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED, justify=tk.LEFT,
        ).pack(side=tk.LEFT, anchor=tk.N, pady=(2, 0))

        # ===== Step 4: 选择对照组 =====
        card4 = self._create_card(scroll_frame, "Step 4: 选择对照组（Control Group）", "🎯")

        ctrl_info = tk.Label(card4, text="对照组用于计算ΔΔCt的基准值，选择Group名称（如有）或样本名称",
                            font=Theme.FONT_SMALL, bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED)
        ctrl_info.pack(anchor=tk.W, pady=(0, 10))

        ctrl_row = tk.Frame(card4, bg=Theme.BG_CARD)
        ctrl_row.pack(fill=tk.X)

        tk.Label(ctrl_row, text="对照组:", font=Theme.FONT_NORMAL,
                bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY).pack(side=tk.LEFT)

        self.dc_ctrl_sample = tk.StringVar(value="")
        self.ctrl_sample_combo = ttk.Combobox(ctrl_row, textvariable=self.dc_ctrl_sample,
                                              font=Theme.FONT_NORMAL, width=25, state="readonly")
        self.ctrl_sample_combo.pack(side=tk.LEFT, padx=(15, 20))
        self.ctrl_sample_combo['values'] = ["请先加载数据并检测"]

        tk.Label(ctrl_row, text="← 选择作为基准的对照组名或样本名", font=Theme.FONT_SMALL,
                bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED).pack(side=tk.LEFT)

        # ===== Step 4.5: 统计与图表设置 (v2.4 新增) =====
        card_stat = self._create_card(scroll_frame, "Step 4·补：定量方法 / 统计检验 / 图表 / Excel 嵌图", "📊")

        method_row = tk.Frame(card_stat, bg=Theme.BG_CARD)
        method_row.pack(fill=tk.X, pady=(0, 8))
        tk.Label(method_row, text="定量方法:", font=Theme.FONT_NORMAL,
                 bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY).pack(side=tk.LEFT)
        self.dc_quant_method = tk.StringVar(value="2^-ΔΔCt")
        ttk.Combobox(
            method_row, textvariable=self.dc_quant_method, state="readonly",
            font=Theme.FONT_NORMAL, width=18,
            values=["2^-ΔΔCt", "Pfaffl (E^-ΔΔCt)"],
        ).pack(side=tk.LEFT, padx=(15, 20))
        ModernButton(
            method_row, text="📊 设置基因效率…", command=self._edit_gene_efficiencies,
            bg=Theme.PRIMARY_LIGHT, width=140, height=32,
        ).pack(side=tk.LEFT, padx=(0, 10))
        # v2.5：每基因扩增效率（默认 2.0）；用户在 Pfaffl 模式下需要这张表
        self.dc_gene_efficiencies: Dict[str, float] = {}
        tk.Label(method_row,
                 text="(选 Pfaffl 时建议为每个基因填扩增效率，例如 1.95 / 100 / 95%)",
                 font=Theme.FONT_SMALL, bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED).pack(side=tk.LEFT)

        stat_row = tk.Frame(card_stat, bg=Theme.BG_CARD)
        stat_row.pack(fill=tk.X, pady=(0, 8))
        tk.Label(stat_row, text="统计检验:", font=Theme.FONT_NORMAL,
                 bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY).pack(side=tk.LEFT)
        self.dc_stat_test = tk.StringVar(value="Student t")
        self.dc_stat_combo = ttk.Combobox(
            stat_row, textvariable=self.dc_stat_test, state="readonly",
            font=Theme.FONT_NORMAL, width=18,
            values=["Student t", "Welch t", "Mann-Whitney U"],
        )
        self.dc_stat_combo.pack(side=tk.LEFT, padx=(15, 20))
        tk.Label(stat_row, text="ΔCt 对照组 vs 处理组；自动 BH-FDR + ns/*/**/***",
                 font=Theme.FONT_SMALL, bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED).pack(side=tk.LEFT)

        fmt_row = tk.Frame(card_stat, bg=Theme.BG_CARD)
        fmt_row.pack(fill=tk.X, pady=(0, 8))
        tk.Label(fmt_row, text="图表格式:", font=Theme.FONT_NORMAL,
                 bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY).pack(side=tk.LEFT)
        self.dc_plot_format = tk.StringVar(value="PNG")
        self.dc_plot_format_combo = ttk.Combobox(
            fmt_row, textvariable=self.dc_plot_format, state="readonly",
            font=Theme.FONT_NORMAL, width=10, values=["PNG", "PDF", "SVG"],
        )
        self.dc_plot_format_combo.pack(side=tk.LEFT, padx=(15, 20))
        self.dc_embed_in_excel = tk.BooleanVar(value=True)
        tk.Checkbutton(
            fmt_row, text="导出 Excel 时嵌入 Fold-change 柱状图",
            variable=self.dc_embed_in_excel, bg=Theme.BG_CARD,
            fg=Theme.TEXT_PRIMARY, activebackground=Theme.BG_CARD,
            font=Theme.FONT_NORMAL,
        ).pack(side=tk.LEFT)

        preset_row = tk.Frame(card_stat, bg=Theme.BG_CARD)
        preset_row.pack(fill=tk.X, pady=(2, 0))
        ModernButton(
            preset_row, text="💾 保存分析预设", command=self._save_analysis_preset,
            bg=Theme.PRIMARY_LIGHT, width=140, height=32,
        ).pack(side=tk.LEFT, padx=(0, 10))
        ModernButton(
            preset_row, text="📂 加载分析预设", command=self._load_analysis_preset,
            bg=Theme.INFO, width=140, height=32,
        ).pack(side=tk.LEFT)
        tk.Label(
            preset_row,
            text="  把「主内参 / 多内参 / 对照组 / 统计方法 / 图表格式 / 嵌图」存为 JSON",
            font=Theme.FONT_SMALL, bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED,
        ).pack(side=tk.LEFT)

        # ===== Step 5: 计算说明 =====
        card5 = self._create_card(scroll_frame, "计算公式说明", "📐")

        formula_text = """计算步骤（使用组内参均值）:
① 组内参Ct均值 = 同一处理组内所有样本的内参基因Ct平均值
② 样本目的基因Ct均值 = 每个样本目的基因所有重复的Ct平均值
③ ΔCt = 样本目的基因Ct均值 - 组内参Ct均值
④ 对照组ΔCt均值 = 对照组所有样本的ΔCt平均值
⑤ ΔΔCt = 样本ΔCt - 对照组ΔCt均值
⑥ 2^(-ΔΔCt) = 相对表达量（Fold Change）

💡 如果没有Group列，则每个样本单独计算内参均值"""

        formula_label = tk.Label(card5, text=formula_text, font=Theme.FONT_SMALL,
                                bg=Theme.BG_CARD, fg=Theme.TEXT_SECONDARY, justify=tk.LEFT)
        formula_label.pack(anchor=tk.W)

        # ===== Step 6: 输出文件名设置卡片 =====
        card6 = self._create_card(scroll_frame, "Step 5: 输出文件名设置", "📝")

        filename_info = tk.Label(card6,
                                text="设置输出文件名模板。支持变量：{date}=日期, {sample}=样本信息, {gene}=基因名",
                                font=Theme.FONT_SMALL, bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED)
        filename_info.pack(anchor=tk.W, pady=(0, 10))

        filename_row = tk.Frame(card6, bg=Theme.BG_CARD)
        filename_row.pack(fill=tk.X)

        tk.Label(filename_row, text="文件名模板:", font=Theme.FONT_NORMAL,
                bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY).pack(side=tk.LEFT)

        self.dc_filename_template = tk.StringVar(value="{date}_{sample}_Analysis and processing")
        filename_entry = tk.Entry(filename_row, textvariable=self.dc_filename_template,
                                 font=Theme.FONT_NORMAL, relief=tk.FLAT,
                                 bg=Theme.BG_INPUT, fg=Theme.TEXT_PRIMARY, width=45,
                                 insertbackground=Theme.PRIMARY)
        filename_entry.pack(side=tk.LEFT, ipady=6, padx=(15, 15))

        tk.Label(filename_row, text=".csv/.xlsx", font=Theme.FONT_SMALL,
                bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED).pack(side=tk.LEFT)

        # 样本信息输入
        sample_info_row = tk.Frame(card6, bg=Theme.BG_CARD)
        sample_info_row.pack(fill=tk.X, pady=(10, 0))

        tk.Label(sample_info_row, text="样本信息 {sample}:", font=Theme.FONT_NORMAL,
                bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY).pack(side=tk.LEFT)

        self.dc_sample_info = tk.StringVar(value="F1-♂-male-hmc-3")
        sample_info_entry = tk.Entry(sample_info_row, textvariable=self.dc_sample_info,
                                    font=Theme.FONT_NORMAL, relief=tk.FLAT,
                                    bg=Theme.BG_INPUT, fg=Theme.TEXT_PRIMARY, width=30,
                                    insertbackground=Theme.PRIMARY)
        sample_info_entry.pack(side=tk.LEFT, ipady=6, padx=(15, 15))

        tk.Label(sample_info_row, text="← 用于文件名中的样本标识", font=Theme.FONT_SMALL,
                bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED).pack(side=tk.LEFT)

        # ===== Step 7: 操作按钮卡片 =====
        card7 = self._create_card(scroll_frame, "Step 6: 执行计算", "🚀")
        btn_frame = tk.Frame(card7, bg=Theme.BG_CARD)
        btn_frame.pack(fill=tk.X, pady=5)

        ModernButton(btn_frame, text="🔢 计算ΔΔCt", command=self._calculate_deltact,
                    bg=Theme.WARNING, width=130, height=42).pack(side=tk.LEFT, padx=(0, 15))
        ModernButton(btn_frame, text="👁️ 预览详细结果", command=self._preview_deltact,
                    bg=Theme.INFO, width=140, height=42).pack(side=tk.LEFT, padx=(0, 15))
        ModernButton(btn_frame, text="📊 预览汇总表", command=self._preview_summary,
                    bg=Theme.PRIMARY_LIGHT, width=130, height=42).pack(side=tk.LEFT, padx=(0, 15))
        ModernButton(btn_frame, text="📋 预览并排格式", command=self._preview_side_by_side,
                    bg="#9333ea", width=140, height=42).pack(side=tk.LEFT)

        # 第二行导出按钮
        btn_frame2 = tk.Frame(card7, bg=Theme.BG_CARD)
        btn_frame2.pack(fill=tk.X, pady=(10, 5))

        ModernButton(btn_frame2, text="💾 导出Excel", command=self._export_deltact,
                    bg=Theme.SUCCESS, width=130, height=42).pack(side=tk.LEFT, padx=(0, 15))
        ModernButton(btn_frame2, text="📄 导出并排格式CSV", command=self._export_side_by_side_csv,
                    bg=Theme.PRIMARY, width=160, height=42).pack(side=tk.LEFT)

        # 第三行：Fold-change 柱状图（需先计算）
        plot_row = tk.Frame(card7, bg=Theme.BG_CARD)
        plot_row.pack(fill=tk.X, pady=(12, 0))
        tk.Label(
            plot_row, text="图表基因:", font=Theme.FONT_NORMAL,
            bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY,
        ).pack(side=tk.LEFT)
        self.dc_plot_gene_var = tk.StringVar(value="(全部)")
        self.dc_plot_gene_combo = ttk.Combobox(
            plot_row, textvariable=self.dc_plot_gene_var,
            font=Theme.FONT_NORMAL, width=18, state="readonly",
        )
        self.dc_plot_gene_combo.pack(side=tk.LEFT, padx=(10, 15))
        self.dc_plot_gene_combo["values"] = ["(全部)"]
        ModernButton(
            plot_row, text="📈 导出 Fold-change 柱状图 (PNG)", command=self._export_dc_foldchange_plot,
            bg="#7c3aed", width=220, height=40,
        ).pack(side=tk.LEFT)

    def _detect_genes_samples(self):
        """检测数据中的基因、样本和分组列表"""
        if self.calculator.raw_data is None:
            messagebox.showwarning("警告", "请先选择并加载数据文件")
            return

        try:
            df = self.calculator.raw_data
            genes, samples = self.calculator.detect_genes_and_samples(
                sample_col=self.dc_sample_col.get(),
                gene_col=self.dc_gene_col.get()
            )

            # 检测Group列
            group_col = self.dc_group_col.get().strip()
            groups = []
            if group_col and group_col in df.columns:
                groups = df[group_col].dropna().unique().tolist()

            if genes:
                self.ref_gene_combo['values'] = genes
                self.ref_gene_combo['state'] = "readonly"
                if len(genes) > 0:
                    # 尝试自动选择常用内参
                    common_refs = ['GAPDH', 'Gapdh', 'gapdh', 'ACTB', 'Actb', 'actb', '18S', 'β-actin', 'HPRT', 'Hprt']
                    selected = genes[0]
                    for ref in common_refs:
                        if ref in genes:
                            selected = ref
                            break
                    self.dc_ref_gene.set(selected)

                # 同步刷新「多内参 (可选)」listbox（v2.4）
                try:
                    self.dc_multi_ref_listbox.delete(0, tk.END)
                    for g in genes:
                        if str(g) != selected:
                            self.dc_multi_ref_listbox.insert(tk.END, str(g))
                except Exception:
                    pass

            # 对照组选择：优先使用Group，否则用Sample
            if groups:
                ctrl_options = groups
            else:
                ctrl_options = samples

            if ctrl_options:
                self.ctrl_sample_combo['values'] = ctrl_options
                self.ctrl_sample_combo['state'] = "readonly"
                if len(ctrl_options) > 0:
                    # 尝试自动选择对照组
                    common_ctrls = ['Control', 'control', 'CTRL', 'ctrl', 'NC', 'WT', 'wt', 'Con', 'CON']
                    selected = ctrl_options[0]
                    for ctrl in common_ctrls:
                        if ctrl in ctrl_options:
                            selected = ctrl
                            break
                    self.dc_ctrl_sample.set(selected)

            if groups:
                self.detect_status_label.config(
                    text=f"✅ 检测到 {len(genes)} 个基因, {len(samples)} 个样本, {len(groups)} 个分组",
                    fg=Theme.SUCCESS
                )
                self.status_var.set(f"✅ 检测完成: {len(genes)} 基因, {len(samples)} 样本, {len(groups)} 分组")
            else:
                self.detect_status_label.config(
                    text=f"✅ 检测到 {len(genes)} 个基因, {len(samples)} 个样本（无分组）",
                    fg=Theme.SUCCESS
                )
                self.status_var.set(f"✅ 检测完成: {len(genes)} 基因, {len(samples)} 样本")

        except Exception as e:
            messagebox.showerror("错误", f"检测失败: {e}\n请确认列名设置正确")
            self.detect_status_label.config(text=f"❌ 检测失败", fg=Theme.ERROR)

    def _preview_summary(self):
        """预览汇总表"""
        if self.calculator.summary is None:
            messagebox.showwarning("警告", "请先执行计算")
            return
        self._show_preview_window("ΔΔCt计算汇总表", self.calculator.summary)

    def _select_deltact_file(self):
        """选择ΔΔCt计算文件"""
        filepath = filedialog.askopenfilename(
            title="选择Ct数据文件",
            filetypes=[
                ("Excel/CSV", "*.xlsx;*.xls;*.csv"),
                ("Excel文件", "*.xlsx;*.xls"),
                ("CSV文件", "*.csv"),
                ("所有文件", "*.*"),
            ],
        )
        if filepath:
            self.deltact_input_file = filepath
            self.deltact_file_label.config(text=f"📄 {Path(filepath).name}",
                                          fg=Theme.TEXT_PRIMARY, bg="#e0f2fe")
            try:
                ext = Path(filepath).suffix.lower()
                if ext in ['.xlsx', '.xls']:
                    # Excel文件：获取所有sheet名称
                    xl = pd.ExcelFile(filepath)
                    sheet_names = xl.sheet_names
                    self.sheet_combo['values'] = sheet_names
                    self.sheet_combo['state'] = "readonly"

                    # 优先选择"长格式"相关的sheet
                    selected_sheet = sheet_names[0]
                    for name in sheet_names:
                        if '长格式' in name or 'long' in name.lower() or 'ΔΔCt' in name:
                            selected_sheet = name
                            break
                    self.dc_sheet_var.set(selected_sheet)

                    # 加载选中的sheet
                    self._load_sheet_data(filepath, selected_sheet)
                    self.status_var.set(f"✅ 已加载Excel，共 {len(sheet_names)} 个Sheet")
                else:
                    # CSV文件：直接加载
                    self.sheet_combo['values'] = ["CSV文件无Sheet"]
                    self.sheet_combo['state'] = "disabled"
                    df = self.calculator.load_data(filepath)
                    self._auto_detect_columns(df)
                    self.status_var.set(f"✅ 已加载 {len(df)} 行数据")
            except Exception as e:
                messagebox.showerror("错误", f"加载失败: {e}")

    def _on_sheet_selected(self, event=None):
        """当用户选择不同的Sheet时"""
        if self.deltact_input_file and self.dc_sheet_var.get():
            self._load_sheet_data(self.deltact_input_file, self.dc_sheet_var.get())

    def _load_sheet_data(self, filepath: str, sheet_name: str):
        """加载指定Sheet的数据"""
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_name)
            self.calculator.raw_data = df
            self._auto_detect_columns(df)
            self.status_var.set(f"✅ 已加载 Sheet '{sheet_name}'，共 {len(df)} 行数据")
        except Exception as e:
            messagebox.showerror("错误", f"加载Sheet失败: {e}")

    def _auto_detect_columns(self, df: pd.DataFrame):
        """自动检测列名并判断数据格式"""
        cols = list(df.columns)

        # 检查是否是长格式（有Sample, Group, Gene, Ct列）
        sample_col = group_col = gene_col = ct_col = None

        for col in cols:
            col_lower = str(col).lower()
            if 'sample' in col_lower or col_lower == 'sample':
                sample_col = col
            elif 'group' in col_lower or '组' in str(col):
                group_col = col
            elif 'gene' in col_lower or 'target' in col_lower:
                gene_col = col
            elif 'ct' in col_lower or 'cp' in col_lower:
                ct_col = col

        if sample_col and gene_col and ct_col:
            self.dc_sample_col.set(sample_col)
            self.dc_gene_col.set(gene_col)
            self.dc_ct_col.set(ct_col)
            if group_col:
                self.dc_group_col.set(group_col)
                self.detect_status_label.config(
                    text="✅ 检测到长格式数据（含Group列），可直接使用",
                    fg=Theme.SUCCESS
                )
            else:
                self.dc_group_col.set("")  # 没有Group列
                self.detect_status_label.config(
                    text="✅ 检测到长格式数据（无Group列，每样本单独计算内参）",
                    fg=Theme.SUCCESS
                )
        else:
            # 可能是宽格式，提示用户
            if len(cols) >= 3:
                self.dc_sample_col.set(cols[0])
                self.dc_gene_col.set(cols[1])
                self.dc_ct_col.set(cols[2])
            self.detect_status_label.config(
                text="⚠️ 可能是宽格式数据，请选择'长格式_ΔΔCt用'的Sheet",
                fg=Theme.WARNING
            )

    def _calculate_deltact(self):
        """执行ΔΔCt计算（支持多内参几何平均 + 统计检验）。"""
        if self.calculator.raw_data is None:
            messagebox.showwarning("警告", "请先选择文件")
            return

        try:
            # 获取Group列名（如果有）
            group_col = self.dc_group_col.get().strip()
            if group_col and group_col in self.calculator.raw_data.columns:
                group_col_param = group_col
            else:
                group_col_param = None

            primary = self.dc_ref_gene.get().strip()
            extras = self._dc_get_multi_refs()
            ref_genes_full: List[str] = list(dict.fromkeys(([primary] if primary else []) + extras))
            stat_method = self._dc_normalize_stat_test()
            quant_method = self._dc_normalize_method()

            df = self.calculator.calculate(
                sample_col=self.dc_sample_col.get(),
                gene_col=self.dc_gene_col.get(),
                ct_col=self.dc_ct_col.get(),
                ref_gene=primary,
                ctrl_sample=self.dc_ctrl_sample.get(),
                group_col=group_col_param,
                ref_genes=ref_genes_full,
                stat_test=stat_method,
                method=quant_method,
                gene_efficiencies=(self.dc_gene_efficiencies if quant_method == "pfaffl" else None),
            )

            mode = "组内参" if group_col_param else "样本内参"
            ref_desc = ref_genes_full[0] if len(ref_genes_full) == 1 else (
                f"{ref_genes_full[0]} + " + ", ".join(ref_genes_full[1:]) + " (几何平均)"
            )
            quant_desc = "Pfaffl (E^-ΔΔCt)" if quant_method == "pfaffl" else "2^-ΔΔCt"
            msg = (f"✅ 计算完成（{mode} | ref={ref_desc} | {stat_method} | "
                   f"{quant_desc}），共 {len(df)} 行")
            self.status_var.set(msg)
            self._refresh_dc_plot_genes()
            messagebox.showinfo(
                "成功",
                f"ΔΔCt 计算完成！\n共 {len(df)} 条记录\n"
                f"内参: {ref_desc}\n统计: {stat_method}\n方法: {quant_desc}"
            )
        except Exception as e:
            messagebox.showerror("错误", f"计算失败: {e}")

    # ---- v2.4 helpers ----
    def _dc_get_multi_refs(self) -> List[str]:
        try:
            sel = self.dc_multi_ref_listbox.curselection()
            return [self.dc_multi_ref_listbox.get(i) for i in sel]
        except Exception:
            return []

    def _dc_normalize_stat_test(self) -> str:
        m = self.dc_stat_test.get().strip().lower()
        if "welch" in m:
            return "welch"
        if "mann" in m or "whitney" in m:
            return "mannwhitney"
        return "ttest"

    def _dc_normalize_method(self) -> str:
        m = self.dc_quant_method.get().strip().lower()
        return "pfaffl" if "pfaffl" in m else "ddct"

    def _edit_gene_efficiencies(self) -> None:
        """弹窗：让用户为每个基因填扩增效率（base 1.0~2.5；也接受百分比 / 0~1）。"""
        genes = list(self.calculator.available_genes) or [
            self.dc_ref_gene.get()
        ] + self._dc_get_multi_refs()
        genes = [g for g in dict.fromkeys(g.strip() for g in genes if g and g.strip())]
        if not genes:
            messagebox.showinfo("提示", "请先「检测基因和样本」生成基因列表。")
            return

        win = tk.Toplevel(self.root)
        win.title("📊 基因扩增效率 (Pfaffl)")
        win.configure(bg=Theme.BG_MAIN)
        win.transient(self.root)
        win.grab_set()
        w, h = 480, min(600, 100 + 36 * len(genes) + 100)
        x = (win.winfo_screenwidth() - w) // 2
        y = (win.winfo_screenheight() - h) // 2
        win.geometry(f"{w}x{h}+{x}+{y}")

        header = tk.Frame(win, bg=Theme.PRIMARY_DARK, height=46)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        tk.Label(header, text="📊 基因扩增效率", font=Theme.FONT_HEADING,
                 bg=Theme.PRIMARY_DARK, fg=Theme.TEXT_WHITE).pack(side=tk.LEFT, padx=18, pady=10)

        body = tk.Frame(win, bg=Theme.BG_CARD)
        body.pack(fill=tk.BOTH, expand=True, padx=16, pady=12)
        tk.Label(body, text=(
            "理论 base = 2.0（100% 扩增效率）。可填:\n"
            "  • 1.95 / 2.05 等 base 值（推荐）\n  • 95 / 100 等百分比\n  • 0.95 / 0.85（相对增量）\n"
            "留空 = 沿用 2.0；选 「2^-ΔΔCt」 时此表会被忽略。"),
                 font=Theme.FONT_SMALL, bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED, justify=tk.LEFT,
                 ).pack(anchor=tk.W, pady=(0, 8))

        canvas = tk.Canvas(body, bg=Theme.BG_CARD, highlightthickness=0)
        scrollbar = ttk.Scrollbar(body, orient="vertical", command=canvas.yview)
        inner = tk.Frame(canvas, bg=Theme.BG_CARD)
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        canvas.create_window((0, 0), window=inner, anchor="nw")
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        entries: Dict[str, tk.StringVar] = {}
        for g in genes:
            row = tk.Frame(inner, bg=Theme.BG_CARD)
            row.pack(fill=tk.X, pady=4)
            tk.Label(row, text=str(g), font=Theme.FONT_NORMAL,
                     bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY, width=22, anchor=tk.W,
                     ).pack(side=tk.LEFT)
            cur = self.dc_gene_efficiencies.get(str(g))
            v = tk.StringVar(value=("" if cur is None else f"{cur:g}"))
            entries[str(g)] = v
            tk.Entry(row, textvariable=v, font=Theme.FONT_NORMAL, relief=tk.FLAT,
                     bg=Theme.BG_INPUT, fg=Theme.TEXT_PRIMARY, width=14,
                     insertbackground=Theme.PRIMARY).pack(side=tk.LEFT, ipady=5, padx=(10, 0))
            tk.Label(row, text="(留空 = 2.0)", font=Theme.FONT_SMALL,
                     bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED).pack(side=tk.LEFT, padx=(8, 0))

        btn_row = tk.Frame(win, bg=Theme.BG_CARD)
        btn_row.pack(fill=tk.X, padx=16, pady=12)

        def on_ok():
            new_eff: Dict[str, float] = {}
            for g, var in entries.items():
                txt = var.get().strip()
                if txt:
                    new_eff[g] = _normalise_efficiency(txt)
            self.dc_gene_efficiencies = new_eff
            n_set = len(new_eff)
            self.status_var.set(f"✅ 已设置 {n_set} 个基因效率（其余按 2.0）")
            win.destroy()

        ModernButton(btn_row, text="✅ 保存", command=on_ok,
                     bg=Theme.SUCCESS, width=120, height=38).pack(side=tk.LEFT, padx=(0, 10))
        ModernButton(btn_row, text="✗ 关闭", command=win.destroy,
                     bg=Theme.TEXT_MUTED, width=100, height=38).pack(side=tk.LEFT)

    def _save_analysis_preset(self) -> None:
        """把 ΔΔCt 标签页当前的「分析配置」存到 presets/。"""
        payload = {
            "version": 1,
            "kind": "analysis",
            "primary_ref_gene": self.dc_ref_gene.get(),
            "multi_ref_genes": self._dc_get_multi_refs(),
            "ctrl_group": self.dc_ctrl_sample.get(),
            "stat_test": self.dc_stat_test.get(),
            "plot_format": self.dc_plot_format.get(),
            "embed_in_excel": bool(self.dc_embed_in_excel.get()),
            "filename_template": self.dc_filename_template.get(),
            "sample_info": self.dc_sample_info.get(),
        }
        path = filedialog.asksaveasfilename(
            title="保存分析预设 (JSON)",
            initialdir=str(self._presets_dir()),
            initialfile="my_analysis_preset.json",
            defaultextension=".json",
            filetypes=[("JSON 预设", "*.json"), ("所有文件", "*.*")],
        )
        if not path:
            return
        try:
            Path(path).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
            self.status_var.set(f"✅ 已保存分析预设: {Path(path).name}")
            messagebox.showinfo("成功", f"分析预设已保存:\n{path}")
        except Exception as exc:
            messagebox.showerror("错误", f"保存失败: {exc}")

    def _load_analysis_preset(self) -> None:
        path = filedialog.askopenfilename(
            title="加载分析预设 (JSON)",
            initialdir=str(self._presets_dir()),
            filetypes=[("JSON 预设", "*.json"), ("所有文件", "*.*")],
        )
        if not path:
            return
        try:
            data = json.loads(Path(path).read_text(encoding="utf-8"))
            if not isinstance(data, dict):
                raise ValueError("预设格式无效")
            if "primary_ref_gene" in data and data["primary_ref_gene"]:
                self.dc_ref_gene.set(str(data["primary_ref_gene"]))
            if "ctrl_group" in data and data["ctrl_group"]:
                self.dc_ctrl_sample.set(str(data["ctrl_group"]))
            if "stat_test" in data and data["stat_test"]:
                self.dc_stat_test.set(str(data["stat_test"]))
            if "plot_format" in data and data["plot_format"]:
                self.dc_plot_format.set(str(data["plot_format"]).upper())
            if "embed_in_excel" in data:
                self.dc_embed_in_excel.set(bool(data["embed_in_excel"]))
            if "filename_template" in data and data["filename_template"]:
                self.dc_filename_template.set(str(data["filename_template"]))
            if "sample_info" in data and data["sample_info"]:
                self.dc_sample_info.set(str(data["sample_info"]))
            multi = data.get("multi_ref_genes") or []
            if multi:
                items = list(self.dc_multi_ref_listbox.get(0, tk.END))
                self.dc_multi_ref_listbox.selection_clear(0, tk.END)
                for g in multi:
                    if g in items:
                        self.dc_multi_ref_listbox.selection_set(items.index(g))
            self.status_var.set(f"✅ 已加载分析预设: {Path(path).name}")
            messagebox.showinfo("成功", f"已加载分析预设:\n{Path(path).name}")
        except Exception as exc:
            messagebox.showerror("错误", f"加载失败: {exc}")

    def _refresh_dc_plot_genes(self) -> None:
        """计算完成后刷新「图表基因」下拉（用于导出 Fold-change 柱状图）。"""
        summ = self.calculator.summary
        if summ is None or "Gene" not in summ.columns:
            self.dc_plot_gene_combo["values"] = ["(全部)"]
            self.dc_plot_gene_var.set("(全部)")
            return
        ref = str(self.dc_ref_gene.get())
        targets = sorted({str(g) for g in summ["Gene"].dropna().unique() if str(g) != ref})
        self.dc_plot_gene_combo["values"] = ["(全部)"] + targets
        self.dc_plot_gene_var.set("(全部)")

    def _export_dc_foldchange_plot(self) -> None:
        """ΔΔCt 标签页：按汇总表导出 2^-ΔΔCt 分组柱状图（均值 ± SEM + 显著性标星）。"""
        if self.calculator.summary is None:
            messagebox.showwarning("警告", "请先执行计算")
            return
        sel = self.dc_plot_gene_var.get().strip()
        genes_filter = None if sel in ("", "(全部)") else [sel]
        fmt = (self.dc_plot_format.get() or "PNG").lower()
        if fmt not in ("png", "pdf", "svg"):
            fmt = "png"
        default_name = f"{datetime.now().strftime('%Y%m%d')}_foldchange_bars.{fmt}"
        filetypes = {
            "png": [("PNG 图片", "*.png"), ("所有文件", "*.*")],
            "pdf": [("PDF 矢量图", "*.pdf"), ("所有文件", "*.*")],
            "svg": [("SVG 矢量图", "*.svg"), ("所有文件", "*.*")],
        }[fmt]
        out = filedialog.asksaveasfilename(
            title="保存 Fold-change 柱状图",
            defaultextension=f".{fmt}",
            initialfile=default_name,
            filetypes=filetypes,
        )
        if not out:
            return
        # 把多内参也排除掉
        ref_extras = self._dc_get_multi_refs()
        ok, msg = export_foldchange_bar_figure(
            self.calculator.summary,
            self.dc_ref_gene.get(),
            out,
            title=f"2^-ΔΔCt (ref={self.dc_ref_gene.get()})",
            genes_filter=genes_filter,
            ctrl_group=self.dc_ctrl_sample.get(),
            ref_genes=ref_extras,
            fig_format=fmt,
        )
        if ok:
            self.status_var.set(f"✅ 已导出图表: {Path(msg).name}")
            messagebox.showinfo("成功", f"已保存:\n{msg}")
        else:
            messagebox.showwarning("无法导出", msg)

    def _preview_deltact(self):
        """预览ΔΔCt结果"""
        if self.calculator.results is None:
            messagebox.showwarning("警告", "请先执行计算")
            return
        self._show_preview_window("ΔΔCt计算结果", self.calculator.results)

    def _preview_side_by_side(self):
        """预览并排格式结果"""
        if self.calculator.side_by_side is None:
            messagebox.showwarning("警告", "请先执行计算")
            return
        self._show_preview_window("并排格式结果（内参-目的基因对比）", self.calculator.side_by_side)

    def _generate_output_filename(self, extension: str = ".csv") -> str:
        """根据模板生成输出文件名"""
        template = self.dc_filename_template.get()
        sample_info = self.dc_sample_info.get()

        # 替换变量
        filename = template.replace("{date}", datetime.now().strftime("%Y%m%d"))
        filename = filename.replace("{sample}", sample_info)

        # 如果有选择的基因，替换{gene}
        if self.dc_ref_gene.get():
            filename = filename.replace("{gene}", self.dc_ref_gene.get())

        return filename + extension

    def _export_side_by_side_csv(self):
        """导出并排格式的CSV文件"""
        if self.calculator.side_by_side is None:
            messagebox.showwarning("警告", "请先执行计算")
            return

        # 生成默认文件名
        default_filename = self._generate_output_filename(".csv")

        output = filedialog.asksaveasfilename(
            title="保存并排格式CSV",
            defaultextension=".csv",
            initialfile=default_filename,
            filetypes=[("CSV文件", "*.csv"), ("所有文件", "*.*")]
        )

        if output:
            self.calculator.side_by_side.to_csv(output, index=False, encoding='utf-8-sig')
            self.status_var.set(f"✅ 已导出: {Path(output).name}")
            messagebox.showinfo("成功", f"已保存并排格式CSV:\n{output}")

    def _export_deltact(self):
        """导出ΔΔCt结果 - 包含详细结果、汇总表和并排格式（v2.4 可嵌入图）。"""
        if self.calculator.results is None:
            messagebox.showwarning("警告", "请先执行计算")
            return

        # 生成默认文件名
        default_filename = self._generate_output_filename(".xlsx")

        output = filedialog.asksaveasfilename(title="保存", defaultextension=".xlsx",
                                             initialfile=default_filename,
                                             filetypes=[("Excel", "*.xlsx")])
        if not output:
            return

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            self.calculator.results.to_excel(writer, sheet_name='详细计算结果', index=False)
            if self.calculator.summary is not None:
                self.calculator.summary.to_excel(writer, sheet_name='汇总表', index=False)
            if self.calculator.side_by_side is not None:
                self.calculator.side_by_side.to_excel(writer, sheet_name='并排格式_Analysis', index=False)
            multi_refs = self._dc_get_multi_refs()
            quant_label = self.dc_quant_method.get()
            eff_str = ", ".join(
                f"{g}={e:g}" for g, e in (self.dc_gene_efficiencies or {}).items()
            ) or "(全部 2.0)"
            params = pd.DataFrame({
                '参数': [
                    '主内参基因', '多内参 (附加)', '统计方法',
                    '定量方法', '基因效率 (Pfaffl)',
                    '对照组样本', '样本列名', '基因列名', 'Ct值列名', '样本信息',
                    '图表格式', '是否嵌入图',
                ],
                '设置值': [
                    self.dc_ref_gene.get(),
                    ', '.join(multi_refs) if multi_refs else '(无)',
                    self.dc_stat_test.get(),
                    quant_label,
                    eff_str if "Pfaffl" in quant_label else "(2^-ΔΔCt 不使用)",
                    self.dc_ctrl_sample.get(),
                    self.dc_sample_col.get(),
                    self.dc_gene_col.get(),
                    self.dc_ct_col.get(),
                    self.dc_sample_info.get(),
                    self.dc_plot_format.get(),
                    'Yes' if self.dc_embed_in_excel.get() else 'No',
                ],
            })
            params.to_excel(writer, sheet_name='计算参数', index=False)

        # 嵌图（v2.4）：导出后再用 openpyxl 单独写一个 Sheet
        embedded = False
        embed_msg = ""
        if self.dc_embed_in_excel.get() and self.calculator.summary is not None:
            try:
                with tempfile.TemporaryDirectory() as td:
                    png_path = Path(td) / "foldchange.png"
                    ok_p, msg_p = export_foldchange_bar_figure(
                        self.calculator.summary,
                        self.dc_ref_gene.get(),
                        png_path,
                        title=f"2^-ΔΔCt (ref={self.dc_ref_gene.get()})",
                        ctrl_group=self.dc_ctrl_sample.get(),
                        ref_genes=multi_refs,
                        fig_format="png",
                    )
                    if ok_p:
                        ok_e, msg_e = embed_image_into_excel(
                            output, png_path, sheet_name="Fold-change 图",
                            note=f"Generated {datetime.now():%Y-%m-%d %H:%M}",
                        )
                        embedded = ok_e
                        embed_msg = msg_e if not ok_e else "嵌入完成"
                    else:
                        embed_msg = msg_p
            except Exception as exc:
                embed_msg = f"嵌图失败: {exc}"

        self.status_var.set(f"✅ 已导出: {Path(output).name}")
        sheets = "• 详细计算结果\n• 汇总表 (含 n / SEM / pvalue / FDR / 显著性)\n• 并排格式_Analysis\n• 计算参数"
        if embedded:
            sheets += "\n• Fold-change 图 (PNG 嵌入)"
        elif self.dc_embed_in_excel.get():
            sheets += f"\n• ⚠ 未嵌入图: {embed_msg}"
        messagebox.showinfo("成功", f"已保存:\n{output}\n\n包含 Sheets:\n{sheets}")

    # ==================== 宽表 → ΔΔCt Tab ====================
    def _create_wide_ddct_tab(self):
        """宽表 → ΔΔCt 标签页

        从「Sample × Gene 宽格式 Ct 表」一步直达「内参/目的基因并排格式」。
        典型输入是「格式转换」标签页导出的 ``原始数据`` 或 ``平均CT值`` Sheet，
        也可手工粘贴的 Excel/CSV。
        """
        tab = tk.Frame(self.notebook, bg=Theme.BG_MAIN)
        self.notebook.add(tab, text="  📊 宽表 → ΔΔCt  ")

        canvas = tk.Canvas(tab, bg=Theme.BG_MAIN, highlightthickness=0)
        scrollbar = ttk.Scrollbar(tab, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg=Theme.BG_MAIN)
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        canvas_win = canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        scroll_frame.bind(
            "<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.bind(
            "<Configure>",
            lambda e: canvas.itemconfig(canvas_win, width=e.width),
        )

        # ===== Step 1: 文件选择 =====
        card1 = self._create_card(scroll_frame, "Step 1: 选择宽格式 Ct 数据 (Excel/CSV)", "📁")
        file_frame = tk.Frame(card1, bg=Theme.BG_CARD)
        file_frame.pack(fill=tk.X)

        self.wide_file_label = tk.Label(
            file_frame, text="  📄 点击右侧按钮选择文件...",
            font=Theme.FONT_NORMAL, bg=Theme.BG_INPUT,
            fg=Theme.TEXT_MUTED, padx=15, pady=12, anchor=tk.W,
        )
        self.wide_file_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

        ModernButton(
            file_frame, text="📂 浏览文件", command=self._select_wide_file,
            bg=Theme.PRIMARY, width=120, height=40,
        ).pack(side=tk.RIGHT, padx=(15, 0))

        sheet_frame = tk.Frame(card1, bg=Theme.BG_CARD)
        sheet_frame.pack(fill=tk.X, pady=(10, 0))

        tk.Label(
            sheet_frame, text="选择 Sheet:", font=Theme.FONT_NORMAL,
            bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY,
        ).pack(side=tk.LEFT)

        self.wide_sheet_var = tk.StringVar(value="")
        self.wide_sheet_combo = ttk.Combobox(
            sheet_frame, textvariable=self.wide_sheet_var,
            font=Theme.FONT_NORMAL, width=25, state="readonly",
        )
        self.wide_sheet_combo.pack(side=tk.LEFT, padx=(15, 15))
        self.wide_sheet_combo['values'] = ["请先选择文件"]
        self.wide_sheet_combo.bind("<<ComboboxSelected>>", self._on_wide_sheet_selected)

        tk.Label(
            sheet_frame,
            text="← 推荐选「原始数据」或「平均CT值」",
            font=Theme.FONT_SMALL, bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED,
        ).pack(side=tk.LEFT)

        # ===== Step 2: 列名 / 检测 =====
        card2 = self._create_card(scroll_frame, "Step 2: 数据识别（自动检测，可手动调整）", "🔧")

        sample_row = tk.Frame(card2, bg=Theme.BG_CARD)
        sample_row.pack(fill=tk.X, pady=8)

        tk.Label(
            sample_row, text="样本/分组列名:", font=Theme.FONT_NORMAL,
            bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY,
        ).pack(side=tk.LEFT)

        self.wide_sample_col = tk.StringVar(value="Sample")
        self.wide_sample_combo = ttk.Combobox(
            sample_row, textvariable=self.wide_sample_col,
            font=Theme.FONT_NORMAL, width=18, state="readonly",
        )
        self.wide_sample_combo.pack(side=tk.LEFT, padx=(15, 15))
        self.wide_sample_combo.bind("<<ComboboxSelected>>", lambda _e: self._refresh_wide_genes_groups())

        tk.Label(
            sample_row,
            text="（选定后，其余数值列自动作为基因）",
            font=Theme.FONT_SMALL, bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED,
        ).pack(side=tk.LEFT)

        self.wide_detect_label = tk.Label(
            card2, text="", font=Theme.FONT_SMALL,
            bg=Theme.BG_CARD, fg=Theme.SUCCESS, justify=tk.LEFT,
        )
        self.wide_detect_label.pack(anchor=tk.W, pady=(8, 0))

        qa_row = tk.Frame(card2, bg=Theme.BG_CARD)
        qa_row.pack(fill=tk.X, pady=(6, 0))
        ModernButton(
            qa_row, text="📉 缺失值统计", command=self._wide_na_report,
            bg=Theme.WARNING, width=130, height=34,
        ).pack(side=tk.LEFT)
        tk.Label(
            qa_row,
            text="  查看各基因列空值/非数值单元格数量，便于质控",
            font=Theme.FONT_SMALL, bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED,
        ).pack(side=tk.LEFT, padx=(10, 0))

        # ===== Step 3: 内参基因 =====
        card3 = self._create_card(scroll_frame, "Step 3: 选择内参基因（Reference Gene）", "🧬")

        ref_row = tk.Frame(card3, bg=Theme.BG_CARD)
        ref_row.pack(fill=tk.X)
        tk.Label(
            ref_row, text="主内参:", font=Theme.FONT_NORMAL,
            bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY,
        ).pack(side=tk.LEFT)

        self.wide_ref_gene = tk.StringVar(value="")
        self.wide_ref_combo = ttk.Combobox(
            ref_row, textvariable=self.wide_ref_gene,
            font=Theme.FONT_NORMAL, width=22, state="readonly",
        )
        self.wide_ref_combo.pack(side=tk.LEFT, padx=(15, 15))
        self.wide_ref_combo['values'] = ["请先加载数据"]
        self.wide_ref_combo.bind("<<ComboboxSelected>>", self._on_wide_ref_changed)

        tk.Label(
            ref_row, text="常用: GAPDH / ACTB / RPL13a / 18S rRNA",
            font=Theme.FONT_SMALL, bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED,
        ).pack(side=tk.LEFT)

        # 多内参（v2.4）：可多选；多选时按几何平均计算综合内参
        wide_multi_row = tk.Frame(card3, bg=Theme.BG_CARD)
        wide_multi_row.pack(fill=tk.X, pady=(10, 0))
        tk.Label(
            wide_multi_row, text="多内参 (可选):", font=Theme.FONT_NORMAL,
            bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY,
        ).pack(side=tk.LEFT, anchor=tk.N)
        wmlb_frame = tk.Frame(wide_multi_row, bg=Theme.BG_CARD)
        wmlb_frame.pack(side=tk.LEFT, padx=(15, 15))
        self.wide_multi_ref_listbox = tk.Listbox(
            wmlb_frame, selectmode=tk.MULTIPLE,
            font=Theme.FONT_NORMAL, height=4, width=22,
            bg=Theme.BG_INPUT, fg=Theme.TEXT_PRIMARY,
            highlightbackground=Theme.BORDER, highlightthickness=1,
            exportselection=False,
        )
        self.wide_multi_ref_listbox.pack(side=tk.LEFT)
        wmlb_sb = ttk.Scrollbar(wmlb_frame, orient="vertical", command=self.wide_multi_ref_listbox.yview)
        wmlb_sb.pack(side=tk.LEFT, fill=tk.Y)
        self.wide_multi_ref_listbox.configure(yscrollcommand=wmlb_sb.set)
        tk.Label(
            wide_multi_row,
            text=("← 不勾选 = 仅用主内参；多选时按几何平均合成综合内参\n"
                  "   不再作为目标基因 (会从目标基因列表中自动剔除)"),
            font=Theme.FONT_SMALL, bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED, justify=tk.LEFT,
        ).pack(side=tk.LEFT, anchor=tk.N, pady=(2, 0))

        # ===== Step 4: 目标基因（多选） =====
        card4 = self._create_card(scroll_frame, "Step 4: 选择目标基因（Target Genes，可多选）", "🎯")

        target_btn_row = tk.Frame(card4, bg=Theme.BG_CARD)
        target_btn_row.pack(fill=tk.X, pady=(0, 8))
        ModernButton(
            target_btn_row, text="✓ 全选", command=lambda: self._toggle_wide_targets(True),
            bg=Theme.INFO, width=85, height=32,
        ).pack(side=tk.LEFT, padx=(0, 8))
        ModernButton(
            target_btn_row, text="✗ 全不选", command=lambda: self._toggle_wide_targets(False),
            bg=Theme.TEXT_MUTED, width=85, height=32,
        ).pack(side=tk.LEFT)

        tk.Label(
            target_btn_row,
            text="  默认勾选所有非内参基因；可手动取消不需要的基因",
            font=Theme.FONT_SMALL, bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED,
        ).pack(side=tk.LEFT, padx=(15, 0))

        # 滚动 checkbox 容器（高度受限，避免一长串基因把整个 tab 顶满）
        check_outer = tk.Frame(card4, bg=Theme.BG_CARD,
                               highlightbackground=Theme.BORDER, highlightthickness=1)
        check_outer.pack(fill=tk.X)
        self.wide_target_canvas = tk.Canvas(
            check_outer, bg=Theme.BG_CARD, height=120,
            highlightthickness=0,
        )
        target_scroll = ttk.Scrollbar(
            check_outer, orient="vertical", command=self.wide_target_canvas.yview,
        )
        self.wide_target_inner = tk.Frame(self.wide_target_canvas, bg=Theme.BG_CARD)
        self.wide_target_canvas.configure(yscrollcommand=target_scroll.set)
        target_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.wide_target_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.wide_target_canvas.create_window(
            (0, 0), window=self.wide_target_inner, anchor="nw",
        )
        self.wide_target_inner.bind(
            "<Configure>",
            lambda e: self.wide_target_canvas.configure(
                scrollregion=self.wide_target_canvas.bbox("all"),
            ),
        )

        self.wide_target_placeholder = tk.Label(
            self.wide_target_inner,
            text="  （先加载数据并选择内参基因，目标基因列表会自动出现）",
            font=Theme.FONT_SMALL, bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED,
        )
        self.wide_target_placeholder.pack(anchor=tk.W, padx=10, pady=10)

        # ===== Step 5: 对照组 =====
        card5 = self._create_card(scroll_frame, "Step 5: 选择对照组（Control Group）", "🎚️")

        ctrl_row = tk.Frame(card5, bg=Theme.BG_CARD)
        ctrl_row.pack(fill=tk.X)
        tk.Label(
            ctrl_row, text="对照组:", font=Theme.FONT_NORMAL,
            bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY,
        ).pack(side=tk.LEFT)

        self.wide_ctrl_group = tk.StringVar(value="")
        self.wide_ctrl_combo = ttk.Combobox(
            ctrl_row, textvariable=self.wide_ctrl_group,
            font=Theme.FONT_NORMAL, width=22, state="readonly",
        )
        self.wide_ctrl_combo.pack(side=tk.LEFT, padx=(15, 15))
        self.wide_ctrl_combo['values'] = ["请先加载数据"]

        tk.Label(
            ctrl_row, text="对照组 ΔCt 均值用作 ΔΔCt 基准",
            font=Theme.FONT_SMALL, bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED,
        ).pack(side=tk.LEFT)

        # ===== Step 6: 文件名模板 =====
        card6 = self._create_card(scroll_frame, "Step 6: 输出文件名设置（可选）", "📝")

        fname_row = tk.Frame(card6, bg=Theme.BG_CARD)
        fname_row.pack(fill=tk.X)

        tk.Label(
            fname_row, text="文件名模板:", font=Theme.FONT_NORMAL,
            bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY,
        ).pack(side=tk.LEFT)

        self.wide_filename_template = tk.StringVar(
            value="{date}_{sample}_Analysis and processing"
        )
        tk.Entry(
            fname_row, textvariable=self.wide_filename_template,
            font=Theme.FONT_NORMAL, relief=tk.FLAT,
            bg=Theme.BG_INPUT, fg=Theme.TEXT_PRIMARY, width=42,
            insertbackground=Theme.PRIMARY,
        ).pack(side=tk.LEFT, ipady=6, padx=(15, 15))
        tk.Label(
            fname_row, text=".csv / .xlsx",
            font=Theme.FONT_SMALL, bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED,
        ).pack(side=tk.LEFT)

        sample_info_row = tk.Frame(card6, bg=Theme.BG_CARD)
        sample_info_row.pack(fill=tk.X, pady=(10, 0))

        tk.Label(
            sample_info_row, text="样本信息 {sample}:", font=Theme.FONT_NORMAL,
            bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY,
        ).pack(side=tk.LEFT)
        self.wide_sample_info = tk.StringVar(value="F1-PFC")
        tk.Entry(
            sample_info_row, textvariable=self.wide_sample_info,
            font=Theme.FONT_NORMAL, relief=tk.FLAT,
            bg=Theme.BG_INPUT, fg=Theme.TEXT_PRIMARY, width=30,
            insertbackground=Theme.PRIMARY,
        ).pack(side=tk.LEFT, ipady=6, padx=(15, 0))

        # ===== Step 7: 操作 =====
        card7 = self._create_card(scroll_frame, "Step 7: 执行计算 & 导出", "🚀")

        btn_row1 = tk.Frame(card7, bg=Theme.BG_CARD)
        btn_row1.pack(fill=tk.X, pady=(0, 8))
        ModernButton(
            btn_row1, text="🔢 计算 ΔΔCt", command=self._calculate_wide_ddct,
            bg=Theme.WARNING, width=130, height=42,
        ).pack(side=tk.LEFT, padx=(0, 12))
        ModernButton(
            btn_row1, text="👁️ 预览长格式", command=self._preview_wide_long,
            bg=Theme.INFO, width=130, height=42,
        ).pack(side=tk.LEFT, padx=(0, 12))
        ModernButton(
            btn_row1, text="📋 预览并排格式", command=self._preview_wide_side,
            bg="#9333ea", width=140, height=42,
        ).pack(side=tk.LEFT, padx=(0, 12))
        ModernButton(
            btn_row1, text="📊 预览汇总表", command=self._preview_wide_summary,
            bg=Theme.PRIMARY_LIGHT, width=130, height=42,
        ).pack(side=tk.LEFT)

        btn_row2 = tk.Frame(card7, bg=Theme.BG_CARD)
        btn_row2.pack(fill=tk.X)
        ModernButton(
            btn_row2, text="📄 导出并排格式 CSV", command=self._export_wide_side_csv,
            bg=Theme.SUCCESS, width=170, height=42,
        ).pack(side=tk.LEFT, padx=(0, 12))
        ModernButton(
            btn_row2, text="💾 导出 Excel (多 Sheet)", command=self._export_wide_excel,
            bg=Theme.PRIMARY, width=185, height=42,
        ).pack(side=tk.LEFT)

        # 统计 / 格式 / 嵌图 (v2.4) + 定量方法 (v2.5)
        wide_method_row = tk.Frame(card7, bg=Theme.BG_CARD)
        wide_method_row.pack(fill=tk.X, pady=(12, 0))
        tk.Label(
            wide_method_row, text="定量方法:", font=Theme.FONT_NORMAL,
            bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY,
        ).pack(side=tk.LEFT)
        self.wide_quant_method = tk.StringVar(value="2^-ΔΔCt")
        ttk.Combobox(
            wide_method_row, textvariable=self.wide_quant_method, state="readonly",
            font=Theme.FONT_NORMAL, width=18,
            values=["2^-ΔΔCt", "Pfaffl (E^-ΔΔCt)"],
        ).pack(side=tk.LEFT, padx=(10, 20))
        ModernButton(
            wide_method_row, text="📊 设置基因效率…", command=self._edit_gene_efficiencies,
            bg=Theme.PRIMARY_LIGHT, width=140, height=32,
        ).pack(side=tk.LEFT, padx=(0, 10))
        tk.Label(
            wide_method_row,
            text="(Pfaffl 时按基因 E^-ΔΔCt；效率表与 ΔΔCt 标签页共享)",
            font=Theme.FONT_SMALL, bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED,
        ).pack(side=tk.LEFT)

        wide_stat_row = tk.Frame(card7, bg=Theme.BG_CARD)
        wide_stat_row.pack(fill=tk.X, pady=(8, 0))
        tk.Label(
            wide_stat_row, text="统计检验:", font=Theme.FONT_NORMAL,
            bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY,
        ).pack(side=tk.LEFT)
        self.wide_stat_test = tk.StringVar(value="Student t")
        ttk.Combobox(
            wide_stat_row, textvariable=self.wide_stat_test, state="readonly",
            font=Theme.FONT_NORMAL, width=18,
            values=["Student t", "Welch t", "Mann-Whitney U"],
        ).pack(side=tk.LEFT, padx=(10, 20))
        tk.Label(
            wide_stat_row, text="图表格式:", font=Theme.FONT_NORMAL,
            bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY,
        ).pack(side=tk.LEFT)
        self.wide_plot_format = tk.StringVar(value="PNG")
        ttk.Combobox(
            wide_stat_row, textvariable=self.wide_plot_format, state="readonly",
            font=Theme.FONT_NORMAL, width=8, values=["PNG", "PDF", "SVG"],
        ).pack(side=tk.LEFT, padx=(10, 20))
        self.wide_embed_in_excel = tk.BooleanVar(value=True)
        tk.Checkbutton(
            wide_stat_row, text="导出 Excel 时嵌入柱状图",
            variable=self.wide_embed_in_excel, bg=Theme.BG_CARD,
            fg=Theme.TEXT_PRIMARY, activebackground=Theme.BG_CARD,
            font=Theme.FONT_NORMAL,
        ).pack(side=tk.LEFT)

        plot_row3 = tk.Frame(card7, bg=Theme.BG_CARD)
        plot_row3.pack(fill=tk.X, pady=(12, 0))
        tk.Label(
            plot_row3, text="图表基因:", font=Theme.FONT_NORMAL,
            bg=Theme.BG_CARD, fg=Theme.TEXT_PRIMARY,
        ).pack(side=tk.LEFT)
        self.wide_plot_gene_var = tk.StringVar(value="(全部)")
        self.wide_plot_gene_combo = ttk.Combobox(
            plot_row3, textvariable=self.wide_plot_gene_var,
            font=Theme.FONT_NORMAL, width=18, state="readonly",
        )
        self.wide_plot_gene_combo.pack(side=tk.LEFT, padx=(10, 15))
        self.wide_plot_gene_combo["values"] = ["(全部)"]
        ModernButton(
            plot_row3, text="📈 导出 Fold-change 柱状图", command=self._export_wide_foldchange_plot,
            bg="#7c3aed", width=200, height=40,
        ).pack(side=tk.LEFT)

    # -------- Wide tab actions --------
    def _select_wide_file(self):
        """选择宽格式 Ct 数据文件。"""
        filepath = filedialog.askopenfilename(
            title="选择宽格式 Ct 数据文件",
            filetypes=[
                ("Excel/CSV", "*.xlsx;*.xls;*.csv"),
                ("Excel 文件", "*.xlsx;*.xls"),
                ("CSV 文件", "*.csv"),
                ("所有文件", "*.*"),
            ],
        )
        if not filepath:
            return
        self.wide_input_file = filepath
        self.wide_file_label.config(
            text=f"📄 {Path(filepath).name}",
            fg=Theme.TEXT_PRIMARY, bg="#e0f2fe",
        )

        try:
            ext = Path(filepath).suffix.lower()
            if ext in {'.xlsx', '.xls'}:
                xl = pd.ExcelFile(filepath)
                sheet_names = xl.sheet_names
                self.wide_sheet_combo['values'] = sheet_names
                self.wide_sheet_combo['state'] = "readonly"
                # 优先选「原始数据」或「平均CT值」（与本软件导出 Sheet 对齐）。
                preferred = sheet_names[0]
                for name in sheet_names:
                    if any(k in name for k in ('原始数据', '平均CT', '平均Ct', 'Sample-Gene')):
                        preferred = name
                        break
                self.wide_sheet_var.set(preferred)
                self._load_wide_sheet(filepath, preferred)
            else:
                self.wide_sheet_combo['values'] = ["CSV 无 Sheet"]
                self.wide_sheet_combo['state'] = "disabled"
                df = read_table_csv(filepath)
                self._populate_wide_dataframe(df)
        except Exception as exc:
            messagebox.showerror("错误", f"加载失败: {exc}")

    def _on_wide_sheet_selected(self, event=None):
        if self.wide_input_file and self.wide_sheet_var.get():
            self._load_wide_sheet(self.wide_input_file, self.wide_sheet_var.get())

    def _load_wide_sheet(self, filepath: str, sheet_name: str):
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_name)
            self._populate_wide_dataframe(df)
        except Exception as exc:
            messagebox.showerror("错误", f"读取 Sheet 失败: {exc}")

    def _populate_wide_dataframe(self, df: pd.DataFrame):
        """加载宽格式 DataFrame 后刷新 UI 候选项。"""
        self.wide_calculator.raw_data = df
        cols = [str(c) for c in df.columns]
        self.wide_sample_combo['values'] = cols
        self.wide_sample_combo['state'] = "readonly"
        # 自动猜测样本列：常见 'Sample' / '样本' / 第一列
        guess = cols[0] if cols else ""
        for c in cols:
            cl = c.lower()
            if cl == 'sample' or '样本' in c or '分组' in c or 'group' == cl:
                guess = c
                break
        self.wide_sample_col.set(guess)
        self._refresh_wide_genes_groups()
        self.status_var.set(f"✅ 已加载宽表，共 {len(df)} 行 × {len(cols)} 列")

    def _refresh_wide_genes_groups(self):
        """根据样本列重算基因列表 / 分组列表，并刷新内参/目标/对照组下拉。"""
        df = self.wide_calculator.raw_data
        if df is None:
            return
        sample_col = self.wide_sample_col.get()
        if sample_col not in df.columns:
            return

        # 所有非样本列里只保留至少含一个数值的列作为「基因列」
        gene_cols: List[str] = []
        for col in df.columns:
            if col == sample_col:
                continue
            numeric = pd.to_numeric(df[col], errors='coerce')
            if numeric.notna().any():
                gene_cols.append(str(col))

        groups = [str(g) for g in df[sample_col].dropna().astype(str).unique() if str(g).strip()]

        # 内参下拉
        common_refs = ['GAPDH', 'Gapdh', 'gapdh', 'ACTB', 'Actb', 'actb',
                       'RPL13a', 'Rpl13a', '18S', 'β-actin', 'HPRT', 'Hprt']
        if gene_cols:
            self.wide_ref_combo['values'] = gene_cols
            self.wide_ref_combo['state'] = "readonly"
            current = self.wide_ref_gene.get()
            if current not in gene_cols:
                pick = gene_cols[0]
                for c in common_refs:
                    if c in gene_cols:
                        pick = c
                        break
                self.wide_ref_gene.set(pick)
        else:
            self.wide_ref_combo['values'] = ["（未检测到数值列）"]

        # 对照组下拉
        common_ctrls = ['Control', 'control', 'CTRL', 'ctrl', 'NC', 'WT', 'wt',
                        'Con', 'CON', 'C']
        if groups:
            self.wide_ctrl_combo['values'] = groups
            self.wide_ctrl_combo['state'] = "readonly"
            current = self.wide_ctrl_group.get()
            if current not in groups:
                pick = groups[0]
                for c in common_ctrls:
                    if c in groups:
                        pick = c
                        break
                self.wide_ctrl_group.set(pick)
        else:
            self.wide_ctrl_combo['values'] = ["（未检测到分组）"]

        self.wide_groups = groups
        self._rebuild_wide_target_checkboxes(gene_cols)
        self._update_wide_detect_label(df, sample_col, gene_cols, groups)

    def _update_wide_detect_label(self, df, sample_col, gene_cols, groups):
        ref = self.wide_ref_gene.get()
        targets_count = sum(1 for g in gene_cols if g != ref)
        rep_summary = ""
        if groups:
            rep_counts = df[sample_col].astype(str).value_counts()
            rep_summary = ", 每组样本数: " + ", ".join(
                f"{g}×{rep_counts.get(g, 0)}" for g in groups[:6]
            )
            if len(groups) > 6:
                rep_summary += " …"
        self.wide_detect_label.config(
            text=(
                f"✅ 检测到 {len(gene_cols)} 个基因列, {len(groups)} 个分组 "
                f"(默认 {targets_count} 个目标基因){rep_summary}"
            )
        )

    def _wide_get_multi_refs(self) -> List[str]:
        try:
            sel = self.wide_multi_ref_listbox.curselection()
            return [self.wide_multi_ref_listbox.get(i) for i in sel]
        except Exception:
            return []

    def _wide_normalize_stat_test(self) -> str:
        m = self.wide_stat_test.get().strip().lower()
        if "welch" in m:
            return "welch"
        if "mann" in m or "whitney" in m:
            return "mannwhitney"
        return "ttest"

    def _wide_normalize_method(self) -> str:
        m = (self.wide_quant_method.get() if hasattr(self, "wide_quant_method") else "").strip().lower()
        return "pfaffl" if "pfaffl" in m else "ddct"

    def _rebuild_wide_target_checkboxes(self, gene_cols: Sequence[str]):
        """根据当前基因列重建「目标基因」多选框区，同时刷新「多内参」listbox。"""
        # 清空现有 widgets / 占位符
        for w in list(self.wide_target_inner.winfo_children()):
            w.destroy()
        self.wide_target_vars.clear()

        ref = self.wide_ref_gene.get()
        # 同步刷新「多内参」listbox（v2.4）
        try:
            self.wide_multi_ref_listbox.delete(0, tk.END)
            for g in gene_cols:
                if str(g) != str(ref):
                    self.wide_multi_ref_listbox.insert(tk.END, str(g))
        except Exception:
            pass

        extras = set(self._wide_get_multi_refs()) if hasattr(self, "wide_multi_ref_listbox") else set()
        targets = [g for g in gene_cols if g != ref and g not in extras]
        if not targets:
            tk.Label(
                self.wide_target_inner,
                text="  （加载数据后选择内参基因，剩余基因会出现在这里）",
                font=Theme.FONT_SMALL, bg=Theme.BG_CARD, fg=Theme.TEXT_MUTED,
            ).pack(anchor=tk.W, padx=10, pady=10)
            return

        per_row = 4
        grid = tk.Frame(self.wide_target_inner, bg=Theme.BG_CARD)
        grid.pack(fill=tk.X, padx=10, pady=8)
        for idx, gene in enumerate(targets):
            var = tk.BooleanVar(value=True)
            self.wide_target_vars[gene] = var
            cb = tk.Checkbutton(
                grid, text=gene, variable=var,
                font=Theme.FONT_NORMAL, bg=Theme.BG_CARD,
                fg=Theme.TEXT_PRIMARY, activebackground=Theme.BG_CARD,
                anchor=tk.W,
            )
            r, c = divmod(idx, per_row)
            cb.grid(row=r, column=c, sticky=tk.W, padx=8, pady=4)

    def _on_wide_ref_changed(self, event=None):
        df = self.wide_calculator.raw_data
        if df is None:
            return
        sample_col = self.wide_sample_col.get()
        gene_cols = [
            str(c) for c in df.columns
            if c != sample_col and pd.to_numeric(df[c], errors='coerce').notna().any()
        ]
        self._rebuild_wide_target_checkboxes(gene_cols)
        groups = [str(g) for g in df[sample_col].dropna().astype(str).unique() if str(g).strip()]
        self._update_wide_detect_label(df, sample_col, gene_cols, groups)

    def _toggle_wide_targets(self, value: bool):
        for var in self.wide_target_vars.values():
            var.set(value)

    def _build_wide_long_df(self) -> Optional[pd.DataFrame]:
        """把当前宽表转换成 calculator 期望的长格式 (Sample, Group, Gene, Ct)。

        样本列的同名分组（如多次出现的 "C"）会按出现顺序自动编号成 ``C_01, C_02, ...``，
        作为 Sample（唯一标识），Group 仍保留原始分组名。
        """
        df = self.wide_calculator.raw_data
        if df is None:
            messagebox.showwarning("警告", "请先选择宽格式数据文件")
            return None
        sample_col = self.wide_sample_col.get()
        if sample_col not in df.columns:
            messagebox.showerror("错误", f"找不到样本列: {sample_col}")
            return None

        ref_gene = self.wide_ref_gene.get()
        extras = self._wide_get_multi_refs()
        all_refs = list(dict.fromkeys(([ref_gene] if ref_gene else []) + extras))
        selected_targets = [g for g, v in self.wide_target_vars.items() if v.get() and g not in all_refs]
        if not ref_gene:
            messagebox.showwarning("警告", "请先选择主内参基因")
            return None
        if not selected_targets:
            messagebox.showwarning("警告", "至少勾选一个目标基因")
            return None
        keep_genes = list(dict.fromkeys(all_refs + selected_targets))

        # 给每行分配唯一 sample_id（{group}_{NN}），保证下游 calculator 能区分
        # 同一分组下的多个生物学重复。
        long_rows = []
        per_group_seen: Dict[str, int] = {}
        for _, row in df.iterrows():
            group_raw = row[sample_col]
            if pd.isna(group_raw):
                continue
            group = str(group_raw).strip()
            if not group:
                continue
            per_group_seen[group] = per_group_seen.get(group, 0) + 1
            sample_id = f"{group}_{per_group_seen[group]:02d}"
            for gene in keep_genes:
                if gene not in df.columns:
                    continue
                val = pd.to_numeric(row[gene], errors='coerce')
                if pd.notna(val):
                    long_rows.append({
                        'Sample': sample_id,
                        'Group': group,
                        'Gene': gene,
                        'Ct': float(val),
                    })

        if not long_rows:
            messagebox.showerror("错误", "没有可用的数值数据，请检查宽表是否含有效 Ct 值")
            return None

        return pd.DataFrame(long_rows)

    def _calculate_wide_ddct(self):
        long_df = self._build_wide_long_df()
        if long_df is None:
            return
        ctrl = self.wide_ctrl_group.get()
        if not ctrl:
            messagebox.showwarning("警告", "请选择对照组")
            return

        try:
            self.wide_calculator.raw_data = long_df  # 让 calculator 直接消费长格式
            extras = self._wide_get_multi_refs()
            primary = self.wide_ref_gene.get().strip()
            ref_full = list(dict.fromkeys(([primary] if primary else []) + extras))
            stat_method = self._wide_normalize_stat_test()
            quant_method = self._wide_normalize_method()
            self.wide_calculator.calculate(
                sample_col='Sample', gene_col='Gene', ct_col='Ct',
                ref_gene=primary, ctrl_sample=ctrl,
                group_col='Group',
                ref_genes=ref_full,
                stat_test=stat_method,
                method=quant_method,
                gene_efficiencies=(self.dc_gene_efficiencies if quant_method == "pfaffl" else None),
            )
            self.wide_long_df = long_df
            n_rows = len(self.wide_calculator.results) if self.wide_calculator.results is not None else 0
            ref_desc = ref_full[0] if len(ref_full) == 1 else (
                f"{ref_full[0]} + {', '.join(ref_full[1:])} (几何平均)"
            )
            self.status_var.set(
                f"✅ 宽表 → ΔΔCt 完成: ref={ref_desc}, "
                f"ctrl={ctrl}, stat={stat_method}, 共 {n_rows} 行结果"
            )
            self._refresh_wide_plot_genes()
            n_targets = sum(
                1 for g, v in self.wide_target_vars.items() if v.get() and g not in ref_full
            )
            messagebox.showinfo(
                "成功",
                f"ΔΔCt 计算完成！\n"
                f"内参: {ref_desc}\n"
                f"目标基因: {n_targets} 个\n"
                f"对照组: {ctrl}\n"
                f"统计: {stat_method}\n"
                f"详细结果 {n_rows} 行，并排格式 {len(self.wide_calculator.side_by_side)} 行",
            )
        except Exception as exc:
            messagebox.showerror("错误", f"计算失败: {exc}")
            # 计算失败时把 raw_data 还原为宽表，避免下次操作出错
            self.wide_calculator.raw_data = self._reload_wide_raw_for_safety()

    def _reload_wide_raw_for_safety(self) -> Optional[pd.DataFrame]:
        if not self.wide_input_file:
            return None
        try:
            ext = Path(self.wide_input_file).suffix.lower()
            if ext in {'.xlsx', '.xls'} and self.wide_sheet_var.get():
                return pd.read_excel(self.wide_input_file, sheet_name=self.wide_sheet_var.get())
            return read_table_csv(self.wide_input_file)
        except Exception:
            return None

    def _preview_wide_long(self):
        if self.wide_long_df is None:
            messagebox.showwarning("警告", "请先执行计算")
            return
        self._show_preview_window("宽表 → 长格式预览", self.wide_long_df)

    def _preview_wide_side(self):
        if self.wide_calculator.side_by_side is None:
            messagebox.showwarning("警告", "请先执行计算")
            return
        self._show_preview_window(
            "宽表 ΔΔCt — 并排格式（内参/目的基因对比）",
            self.wide_calculator.side_by_side,
        )

    def _preview_wide_summary(self):
        if self.wide_calculator.summary is None:
            messagebox.showwarning("警告", "请先执行计算")
            return
        self._show_preview_window("宽表 ΔΔCt — 汇总表", self.wide_calculator.summary)

    def _wide_default_filename(self, ext: str) -> str:
        template = self.wide_filename_template.get()
        sample_info = self.wide_sample_info.get()
        ref = self.wide_ref_gene.get()
        out = template.replace("{date}", datetime.now().strftime("%Y%m%d"))
        out = out.replace("{sample}", sample_info)
        out = out.replace("{gene}", ref)
        return out + ext

    def _export_wide_side_csv(self):
        if self.wide_calculator.side_by_side is None:
            messagebox.showwarning("警告", "请先执行计算")
            return
        output = filedialog.asksaveasfilename(
            title="保存并排格式 CSV",
            defaultextension=".csv",
            initialfile=self._wide_default_filename(".csv"),
            filetypes=[("CSV 文件", "*.csv"), ("所有文件", "*.*")],
        )
        if not output:
            return
        self.wide_calculator.side_by_side.to_csv(output, index=False, encoding='utf-8-sig')
        self.status_var.set(f"✅ 已导出: {Path(output).name}")
        messagebox.showinfo("成功", f"已保存并排格式 CSV:\n{output}")

    def _export_wide_excel(self):
        if self.wide_calculator.results is None:
            messagebox.showwarning("警告", "请先执行计算")
            return
        output = filedialog.asksaveasfilename(
            title="保存 Excel",
            defaultextension=".xlsx",
            initialfile=self._wide_default_filename(".xlsx"),
            filetypes=[("Excel", "*.xlsx")],
        )
        if not output:
            return
        extras = self._wide_get_multi_refs()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            if self.wide_long_df is not None:
                self.wide_long_df.to_excel(writer, sheet_name='长格式_来自宽表', index=False)
            self.wide_calculator.results.to_excel(writer, sheet_name='详细计算结果', index=False)
            if self.wide_calculator.summary is not None:
                self.wide_calculator.summary.to_excel(writer, sheet_name='汇总表', index=False)
            if self.wide_calculator.side_by_side is not None:
                self.wide_calculator.side_by_side.to_excel(
                    writer, sheet_name='并排格式_Analysis', index=False,
                )
            quant_label = self.wide_quant_method.get() if hasattr(self, "wide_quant_method") else "2^-ΔΔCt"
            eff_str = ", ".join(
                f"{g}={e:g}" for g, e in (self.dc_gene_efficiencies or {}).items()
            ) or "(全部 2.0)"
            params = pd.DataFrame({
                '参数': [
                    '模式', '主内参基因', '多内参 (附加)', '统计方法',
                    '定量方法', '基因效率 (Pfaffl)',
                    '对照组', '样本信息', '目标基因(选中)', '图表格式', '是否嵌入图',
                ],
                '设置值': [
                    '宽表 → ΔΔCt',
                    self.wide_ref_gene.get(),
                    ', '.join(extras) if extras else '(无)',
                    self.wide_stat_test.get(),
                    quant_label,
                    eff_str if "Pfaffl" in quant_label else "(2^-ΔΔCt 不使用)",
                    self.wide_ctrl_group.get(),
                    self.wide_sample_info.get(),
                    ', '.join(g for g, v in self.wide_target_vars.items()
                              if v.get() and g != self.wide_ref_gene.get() and g not in extras),
                    self.wide_plot_format.get(),
                    'Yes' if self.wide_embed_in_excel.get() else 'No',
                ],
            })
            params.to_excel(writer, sheet_name='计算参数', index=False)

        # 嵌图
        embedded = False
        embed_msg = ""
        if self.wide_embed_in_excel.get() and self.wide_calculator.summary is not None:
            try:
                with tempfile.TemporaryDirectory() as td:
                    png_path = Path(td) / "wide_foldchange.png"
                    ok_p, msg_p = export_foldchange_bar_figure(
                        self.wide_calculator.summary,
                        self.wide_ref_gene.get(),
                        png_path,
                        title=f"2^-ΔΔCt — 宽表 (ref={self.wide_ref_gene.get()})",
                        ctrl_group=self.wide_ctrl_group.get(),
                        ref_genes=extras,
                        fig_format="png",
                    )
                    if ok_p:
                        ok_e, msg_e = embed_image_into_excel(
                            output, png_path, sheet_name="Fold-change 图",
                            note=f"Generated {datetime.now():%Y-%m-%d %H:%M}",
                        )
                        embedded = ok_e
                        embed_msg = "嵌入完成" if ok_e else msg_e
                    else:
                        embed_msg = msg_p
            except Exception as exc:
                embed_msg = f"嵌图失败: {exc}"

        self.status_var.set(f"✅ 已导出: {Path(output).name}")
        sheets = "• 长格式_来自宽表\n• 详细计算结果\n• 汇总表 (含 n / SEM / pvalue / FDR / 显著性)\n• 并排格式_Analysis\n• 计算参数"
        if embedded:
            sheets += "\n• Fold-change 图 (PNG 嵌入)"
        elif self.wide_embed_in_excel.get():
            sheets += f"\n• ⚠ 未嵌入图: {embed_msg}"
        messagebox.showinfo("成功", f"已保存:\n{output}\n\n包含 Sheets:\n{sheets}")

    def _wide_na_report(self) -> None:
        """宽表各基因列缺失 / 非数值统计。"""
        df = self.wide_calculator.raw_data
        if df is None:
            messagebox.showwarning("警告", "请先加载宽表数据")
            return
        sample_col = self.wide_sample_col.get()
        if sample_col not in df.columns:
            messagebox.showerror("错误", "样本列无效")
            return
        rows: List[Dict] = []
        for col in df.columns:
            if col == sample_col:
                continue
            ser = df[col]
            numeric = pd.to_numeric(ser, errors="coerce")
            n_na = int(numeric.isna().sum())
            n_ok = int(numeric.notna().sum())
            rows.append({"基因列": str(col), "有效数值": n_ok, "空或非数": n_na})
        rep = pd.DataFrame(rows)
        self._show_preview_window(
            f"缺失值统计（宽表 n={len(df)}，样本列={sample_col}）",
            rep,
        )

    def _refresh_wide_plot_genes(self) -> None:
        """宽表 ΔΔCt 计算完成后刷新图表基因下拉。"""
        summ = self.wide_calculator.summary
        if summ is None or "Gene" not in summ.columns:
            self.wide_plot_gene_combo["values"] = ["(全部)"]
            self.wide_plot_gene_var.set("(全部)")
            return
        ref = str(self.wide_ref_gene.get())
        selected = [g for g, v in self.wide_target_vars.items() if v.get()]
        targets = sorted({str(g) for g in summ["Gene"].dropna().unique() if str(g) != ref and g in selected})
        self.wide_plot_gene_combo["values"] = ["(全部)"] + targets
        self.wide_plot_gene_var.set("(全部)")

    def _export_wide_foldchange_plot(self) -> None:
        """宽表标签页：导出 Fold-change 柱状图（PNG / PDF / SVG + 显著性标星）。"""
        if self.wide_calculator.summary is None:
            messagebox.showwarning("警告", "请先执行计算")
            return
        sel = self.wide_plot_gene_var.get().strip()
        genes_filter = None if sel in ("", "(全部)") else [sel]
        fmt = (self.wide_plot_format.get() or "PNG").lower()
        if fmt not in ("png", "pdf", "svg"):
            fmt = "png"
        default_name = f"{datetime.now().strftime('%Y%m%d')}_wide_foldchange_bars.{fmt}"
        filetypes = {
            "png": [("PNG 图片", "*.png"), ("所有文件", "*.*")],
            "pdf": [("PDF 矢量图", "*.pdf"), ("所有文件", "*.*")],
            "svg": [("SVG 矢量图", "*.svg"), ("所有文件", "*.*")],
        }[fmt]
        out = filedialog.asksaveasfilename(
            title="保存 Fold-change 柱状图",
            defaultextension=f".{fmt}",
            initialfile=default_name,
            filetypes=filetypes,
        )
        if not out:
            return
        extras = self._wide_get_multi_refs()
        ok, msg = export_foldchange_bar_figure(
            self.wide_calculator.summary,
            self.wide_ref_gene.get(),
            out,
            title=f"2^-ΔΔCt — 宽表 (ref={self.wide_ref_gene.get()})",
            genes_filter=genes_filter,
            ctrl_group=self.wide_ctrl_group.get(),
            ref_genes=extras,
            fig_format=fmt,
        )
        if ok:
            self.status_var.set(f"✅ 已导出图表: {Path(msg).name}")
            messagebox.showinfo("成功", f"已保存:\n{msg}")
        else:
            messagebox.showwarning("无法导出", msg)

    # ==================== 通用方法 ====================
    def _show_preview_window(self, title: str, df: pd.DataFrame):
        """显示预览窗口 - 美化版"""
        win = tk.Toplevel(self.root)
        win.title(f"📊 {title}")
        win.geometry("1250x650")
        win.configure(bg=Theme.BG_MAIN)

        # 居中显示
        win.update_idletasks()
        w, h = 1250, 650
        x = (win.winfo_screenwidth() - w) // 2
        y = (win.winfo_screenheight() - h) // 2
        win.geometry(f"{w}x{h}+{x}+{y}")

        # 顶部标题栏
        header = tk.Frame(win, bg=Theme.PRIMARY, height=50)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        tk.Label(header, text=f"📊 {title}", font=Theme.FONT_HEADING,
                bg=Theme.PRIMARY, fg=Theme.TEXT_WHITE).pack(side=tk.LEFT, padx=20, pady=12)
        tk.Label(header, text=f"共 {len(df)} 行数据", font=Theme.FONT_SMALL,
                bg=Theme.PRIMARY, fg="#c7d2fe").pack(side=tk.RIGHT, padx=20)

        # 表格区域
        frame = tk.Frame(win, bg=Theme.BG_CARD)
        frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)

        # Treeview - 交替行颜色
        tree = ttk.Treeview(frame, columns=list(df.columns), show='headings')
        tree.tag_configure('oddrow', background='#f8fafc')
        tree.tag_configure('evenrow', background='#ffffff')

        for col in df.columns:
            tree.heading(col, text=col)
            # 根据列名设置宽度
            col_width = max(100, len(str(col)) * 12)
            tree.column(col, width=col_width, anchor=tk.CENTER, minwidth=80)

        for idx, row in df.head(500).iterrows():
            values = ['' if pd.isna(v) else (f"{v:.4f}" if isinstance(v, float) else str(v))
                     for v in row]
            tag = 'oddrow' if idx % 2 else 'evenrow'
            tree.insert('', 'end', values=values, tags=(tag,))

        vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        # 底部操作栏
        bottom = tk.Frame(win, bg=Theme.BG_CARD, height=50)
        bottom.pack(fill=tk.X, side=tk.BOTTOM)
        bottom.pack_propagate(False)

        tk.Frame(bottom, bg=Theme.BORDER, height=1).pack(fill=tk.X)

        btn_frame = tk.Frame(bottom, bg=Theme.BG_CARD)
        btn_frame.pack(pady=8)

        # 复制到剪贴板按钮
        def copy_to_clipboard():
            win.clipboard_clear()
            win.clipboard_append(df.to_csv(sep='\t', index=False))
            self.status_var.set("📋 已复制到剪贴板")

        ModernButton(btn_frame, text="📋 复制全部", command=copy_to_clipboard,
                    bg=Theme.PRIMARY, width=120, height=34).pack(side=tk.LEFT, padx=10)
        ModernButton(btn_frame, text="✖ 关闭", command=win.destroy,
                    bg=Theme.TEXT_SECONDARY, width=100, height=34).pack(side=tk.LEFT)

    def run(self):
        """运行"""
        self.root.mainloop()


def main():
    app = CompleteGUI()
    app.run()


if __name__ == '__main__':
    main()

